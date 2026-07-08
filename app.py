from flask import Flask, abort, flash, g, redirect, render_template, request, send_file, session, url_for
import calendar
import hmac
import json
import os
import secrets
import sys
import re
import sqlite3
import threading
import time
from datetime import datetime, timedelta, timezone
from io import BytesIO
from urllib.parse import quote, urlparse
from urllib.request import urlopen, Request
from urllib.error import URLError, HTTPError

from openpyxl import Workbook, load_workbook
from werkzeug.security import check_password_hash, generate_password_hash

import integrations
import telegram_bot

def _sqlite_db_path():
    raw = os.environ.get("DATABASE_PATH") or os.environ.get("VB_DATABASE_PATH") or "database.db"
    return os.path.abspath(raw)


DB_PATH = _sqlite_db_path()


def _prepare_sqlite_storage():
    """Ensure parent dir exists; catch Docker bind-mount mistakes (path is a directory)."""
    parent = os.path.dirname(DB_PATH)
    if parent and not os.path.isdir(parent):
        try:
            os.makedirs(parent, mode=0o755, exist_ok=True)
        except OSError as exc:
            raise RuntimeError(
                f"Cannot create database directory {parent!r}: {exc}"
            ) from exc

    if os.path.exists(DB_PATH) and os.path.isdir(DB_PATH):
        raise RuntimeError(
            f"DATABASE_PATH {DB_PATH!r} is a directory, not a SQLite file. "
            "Docker often creates a directory when a single-file bind mount source was missing."
        ) from None


EXPORT_FORMAT_VERSION = 1
LIST_PAGE_SIZE = 75
TRANSFER_LOG_LIMIT = 10
ALLOW_REGISTRATION = os.environ.get("ALLOW_REGISTRATION", "true").lower() in ("1", "true", "yes")
_USERNAME_RE = re.compile(r"^[a-zA-Z0-9_.-]{3,32}$")
# Precomputed hash of a random value, compared against when a username is not
# found so login timing stays constant regardless of username existence.
_DUMMY_PASSWORD_HASH = generate_password_hash(secrets.token_urlsafe(16))

SHEET_META = "_meta"
SHEET_ACCOUNTS = "Accounts"
SHEET_EXPENSE_CATEGORIES = "ExpenseCategories"
SHEET_INCOME_CATEGORIES = "IncomeCategories"
SHEET_EXPENSES = "Expenses"
SHEET_INCOME = "Income"
ALLOWED_PANELS = {
    "home",
    "expenses",
    "income",
    "recurring",
    "transfer",
    "summary",
    "yearly",
    "reports",
    "investments",
    "settings",
}
SETTINGS_SECTIONS = {"general", "banks", "expenses", "income", "export", "migration", "integrations"}
INVESTMENTS_SECTIONS = {"crypto", "stocks"}
REPORTS_SECTIONS = {"bank", "crypto", "stocks"}
FINNHUB_API_KEY = os.environ.get("FINNHUB_API_KEY", "").strip()

def _env_flag(name: str, default: bool) -> bool:
    raw = os.environ.get(name)
    if raw is None:
        return default
    return raw.strip().lower() in ("1", "true", "yes", "on")


app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev-change-me")
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    # Enable only when served over HTTPS (set VB_SECURE_COOKIES=true behind TLS).
    SESSION_COOKIE_SECURE=_env_flag("VB_SECURE_COOKIES", False),
)

CSRF_FIELD_NAME = "_csrf_token"  # noqa: S105 - form field name, not a secret
_CSRF_SAFE_METHODS = frozenset({"GET", "HEAD", "OPTIONS", "TRACE"})
_CSRF_EXEMPT_ENDPOINTS = frozenset({"static"})


def get_csrf_token() -> str:
    token = session.get("_csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        session["_csrf_token"] = token
    return token


@app.context_processor
def _inject_csrf_token():
    return {"csrf_token": get_csrf_token()}


@app.before_request
def _csrf_protect():
    if request.method in _CSRF_SAFE_METHODS:
        return None
    if request.endpoint in _CSRF_EXEMPT_ENDPOINTS:
        return None
    expected = session.get("_csrf_token")
    submitted = request.form.get(CSRF_FIELD_NAME) or request.headers.get("X-CSRF-Token", "")
    if not expected or not submitted or not hmac.compare_digest(str(expected), str(submitted)):
        abort(400, description="Invalid or missing CSRF token.")
    return None


@app.before_request
def _require_login():
    if request.endpoint in ("login", "register", "static", None):
        return None
    uid = session.get("user_id")
    if not uid:
        return redirect(url_for("login", next=request.path))
    try:
        g.user_id = int(uid)
    except (TypeError, ValueError):
        session.pop("user_id", None)
        session.pop("username", None)
        return redirect(url_for("login", next=request.path))
    g.username = session.get("username") or ""
    return None


def normalize_expense_amount(raw):
    """Signed expense movements: negative = spending, positive = refund/credit."""
    try:
        value = float(str(raw).strip())
    except (TypeError, ValueError):
        value = 0.0
    return value


def normalize_income_amount(raw):
    """Income amounts are stored positive."""
    try:
        value = float(str(raw).strip())
    except (TypeError, ValueError):
        value = 0.0
    return abs(value)


def normalize_list_page(value):
    try:
        return max(1, int(str(value).strip()))
    except (TypeError, ValueError):
        return 1


def normalize_optional_category_id(raw):
    """Positive category id from query/form string, or None."""
    s = (raw or "").strip()
    if not s:
        return None
    try:
        n = int(s)
        return n if n > 0 else None
    except (TypeError, ValueError):
        return None


def normalize_day_of_month(value):
    try:
        d = int(str(value).strip())
        return min(31, max(1, d))
    except (TypeError, ValueError):
        return 1


def _parse_row_created_date(row):
    raw = row["created_at"]
    if raw is None:
        return datetime.now().date()
    s = str(raw)
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").date()
    except ValueError:
        return datetime.now().date()


def _due_date_in_month(ym: str, day_of_month: int):
    y, m = map(int, ym.split("-", 1))
    last = calendar.monthrange(y, m)[1]
    d = min(day_of_month, last)
    return datetime(y, m, d).date()


def _month_iter(start_ym: str, end_ym: str):
    y, m = map(int, start_ym.split("-", 1))
    ey, em = map(int, end_ym.split("-", 1))
    while (y, m) <= (ey, em):
        yield f"{y:04d}-{m:02d}"
        m += 1
        if m > 12:
            m = 1
            y += 1


def apply_recurring_entries(conn, user_id):
    """Insert expense/income rows for due recurring rules (catch-up through current month)."""
    uid = int(user_id)
    today = datetime.now().date()
    today_ym = f"{today.year:04d}-{today.month:02d}"
    rules = conn.execute(
        """
        SELECT id, entry_type, amount, category_id, account_id, day_of_month, notes, created_at
        FROM recurring_entries
        WHERE enabled = 1 AND user_id = ?
        """,
        (uid,),
    ).fetchall()
    posted = 0
    for rule in rules:
        created_d = _parse_row_created_date(rule)
        start_ym = f"{created_d.year:04d}-{created_d.month:02d}"
        note_text = (rule["notes"] or "").strip()
        prefix = "[Recurring] "
        full_notes = f"{prefix}{note_text}" if note_text else prefix.strip()

        for ym in _month_iter(start_ym, today_ym):
            exists = conn.execute(
                "SELECT 1 FROM recurring_applied WHERE recurring_id = ? AND ym = ?",
                (rule["id"], ym),
            ).fetchone()
            if exists:
                continue
            due = _due_date_in_month(ym, rule["day_of_month"])
            if due < created_d:
                continue
            if today < due:
                continue
            et = rule["entry_type"]
            if et == "expense":
                conn.execute(
                    """
                    INSERT INTO expenses (user_id, notes, amount, category_id, account_id, spent_at)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        uid,
                        full_notes,
                        -abs(float(rule["amount"])),
                        int(rule["category_id"]),
                        int(rule["account_id"]),
                        due.isoformat(),
                    ),
                )
            elif et == "income":
                conn.execute(
                    """
                    INSERT INTO income_entries (user_id, notes, amount, category_id, account_id, received_at)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        uid,
                        full_notes,
                        abs(float(rule["amount"])),
                        int(rule["category_id"]),
                        int(rule["account_id"]),
                        due.isoformat(),
                    ),
                )
            else:
                continue
            conn.execute(
                "INSERT INTO recurring_applied (recurring_id, ym) VALUES (?, ?)",
                (rule["id"], ym),
            )
            posted += 1
    return posted


def normalize_year(value):
    raw = (value or "").strip()
    if not raw:
        return datetime.now().year
    try:
        year_num = int(raw)
        if year_num < 2000 or year_num > 2100:
            raise ValueError
        return year_num
    except ValueError:
        return datetime.now().year


def resolve_month_filter_from_request():
    """Primary UI month: ?month=YYYY-MM or ?cal_year=&cal_month= (from dropdowns)."""
    raw = (request.args.get("month") or "").strip()
    if raw:
        return normalize_month(raw)
    y_raw = (request.args.get("cal_year") or "").strip()
    m_raw = (request.args.get("cal_month") or "").strip()
    if y_raw and m_raw:
        try:
            y, m = int(y_raw), int(m_raw)
            if 2000 <= y <= 2100 and 1 <= m <= 12:
                return f"{y:04d}-{m:02d}"
        except ValueError:
            pass
    return normalize_month("")


def resolve_list_month_filter(legacy_key, year_key, month_key):
    """Expense/income list month: legacy ?exp_month= or dropdown ?exp_cal_year=&exp_cal_month=."""
    m_raw = (request.args.get(month_key) or "").strip()
    if not m_raw:
        return parse_optional_month(request.args.get(legacy_key))
    y_raw = (request.args.get(year_key) or "").strip()
    if not y_raw:
        return parse_optional_month(request.args.get(legacy_key))
    try:
        formed = f"{int(y_raw):04d}-{int(m_raw):02d}"
        return parse_optional_month(formed)
    except ValueError:
        return parse_optional_month(request.args.get(legacy_key))


def parse_optional_month(value):
    """YYYY-MM for list filters, or None when absent / invalid."""
    raw = (value or "").strip()
    if not raw:
        return None
    try:
        parts = raw.split("-", 1)
        if len(parts) != 2:
            return None
        year_num, month_num = int(parts[0]), int(parts[1])
        if year_num < 2000 or year_num > 2100 or month_num < 1 or month_num > 12:
            return None
        return f"{year_num:04d}-{month_num:02d}"
    except ValueError:
        return None


def month_bounds_dates(ym_str):
    """Half-open calendar month as YYYY-MM-DD dates (start inclusive, end exclusive)."""
    ys, ms = ym_str.split("-", 1)
    month_start = datetime(int(ys), int(ms), 1)
    month_end = month_start + timedelta(days=32)
    month_end = month_end.replace(day=1)
    return (month_start.strftime("%Y-%m-%d"), month_end.strftime("%Y-%m-%d"))


def normalize_month(value):
    raw = (value or "").strip()
    if not raw:
        return datetime.now().strftime("%Y-%m")
    try:
        year_str, month_str = raw.split("-", 1)
        month_year = int(year_str)
        month_num = int(month_str)
        if month_num < 1 or month_num > 12:
            raise ValueError
        return f"{month_year:04d}-{month_num:02d}"
    except ValueError:
        return datetime.now().strftime("%Y-%m")


def normalize_settings_section(value):
    section = (value or "").strip().lower()
    return section if section in SETTINGS_SECTIONS else "general"


def coerce_txn_day(raw):
    """Normalize stored/form raw values to YYYY-MM-DD, or '' if empty."""
    if raw is None:
        return ""
    text = str(raw).strip()
    if not text:
        return ""
    if len(text) >= 10 and text[4] == "-" and text[7] == "-":
        head = text[:10]
        try:
            datetime.strptime(head, "%Y-%m-%d")
            return head
        except ValueError:
            pass
    try:
        normalized = text.replace(" ", "T", 1)
        if len(normalized) == 10:
            normalized = f"{normalized}T00:00:00"
        parsed = datetime.fromisoformat(normalized)
        return parsed.date().isoformat()
    except ValueError:
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                continue
    return text[:10] if len(text) >= 10 else ""


def normalize_txn_day_from_form(raw):
    """Form submission: blank date defaults to today (calendar day only)."""
    return coerce_txn_day(raw) or datetime.now().date().isoformat()


@app.template_filter("txn_day")
def txn_day_filter(raw):
    return coerce_txn_day(raw)


def get_connection():
    try:
        conn = sqlite3.connect(DB_PATH, timeout=30.0)
    except sqlite3.OperationalError as exc:
        raise RuntimeError(
            f"SQLite cannot open {DB_PATH!r}: {exc}. "
            "Prefer mounting a folder (not one file), especially if the host path is on SMB/NFS: "
            "volumes: ['./budget-data:/app/data'] and env DATABASE_PATH=/app/data/database.db"
        ) from exc
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def fetch_account_balances_through(conn, balance_cutoff_d, user_id):
    """Per-account balance with income, expenses, and transfers strictly before balance_cutoff_d (YYYY-MM-DD)."""
    uid = int(user_id)
    return conn.execute(
        """
        SELECT
            a.id,
            a.name,
            a.opening_balance
                + COALESCE(income_totals.total_income, 0)
                + COALESCE(expense_totals.total_expenses, 0)
                + COALESCE(transfer_in.t_in, 0)
                - COALESCE(transfer_out.t_out, 0) AS current_balance
        FROM accounts a
        LEFT JOIN (
            SELECT account_id, SUM(amount) AS total_income
            FROM income_entries
            WHERE user_id = ? AND date(received_at) < date(?)
            GROUP BY account_id
        ) income_totals ON income_totals.account_id = a.id
        LEFT JOIN (
            SELECT account_id, SUM(amount) AS total_expenses
            FROM expenses
            WHERE user_id = ? AND date(spent_at) < date(?)
            GROUP BY account_id
        ) expense_totals ON expense_totals.account_id = a.id
        LEFT JOIN (
            SELECT to_account_id AS account_id, SUM(amount) AS t_in
            FROM account_transfers
            WHERE user_id = ? AND date(transferred_at) < date(?)
            GROUP BY to_account_id
        ) transfer_in ON transfer_in.account_id = a.id
        LEFT JOIN (
            SELECT from_account_id AS account_id, SUM(amount) AS t_out
            FROM account_transfers
            WHERE user_id = ? AND date(transferred_at) < date(?)
            GROUP BY from_account_id
        ) transfer_out ON transfer_out.account_id = a.id
        WHERE a.user_id = ?
        ORDER BY a.name
        """,
        (
            uid,
            balance_cutoff_d,
            uid,
            balance_cutoff_d,
            uid,
            balance_cutoff_d,
            uid,
            balance_cutoff_d,
            uid,
        ),
    ).fetchall()


def month_snap_cutoff(year: int, month_num: int, today_d) -> str:
    """Exclusive date cutoff for balance/portfolio snapshots (matches bank balance chart)."""
    month_start = datetime(year, month_num, 1)
    month_end_exclusive = month_start + timedelta(days=32)
    month_end_exclusive = month_end_exclusive.replace(day=1)
    month_end_cutoff = month_end_exclusive.strftime("%Y-%m-%d")
    if year == today_d.year and month_num == today_d.month:
        snap = (today_d + timedelta(days=1)).isoformat()
    else:
        snap = month_end_cutoff
    # Never snapshot after today (future months in the selected year).
    cap = (today_d + timedelta(days=1)).isoformat()
    if snap > cap:
        return cap
    return snap


def account_balance_at_cutoff(
    conn, balance_cutoff_d: str, user_id, account_id: int | None = None
) -> float:
    """Single-account or all-accounts balance strictly before balance_cutoff_d."""
    rows = fetch_account_balances_through(conn, balance_cutoff_d, user_id)
    if account_id is None:
        return sum(float(r["current_balance"]) for r in rows)
    aid = int(account_id)
    for r in rows:
        if int(r["id"]) == aid:
            return float(r["current_balance"])
    return 0.0


def monthly_total_balances_for_year(
    conn, year: int, today_d, user_id, account_id: int | None = None
) -> list[float]:
    """Account balance(s) after each calendar month; optional single-account filter."""
    balances = []
    for month_num in range(1, 13):
        snap_cutoff = month_snap_cutoff(year, month_num, today_d)
        balances.append(
            account_balance_at_cutoff(conn, snap_cutoff, user_id, account_id)
        )
    return balances


def build_monthly_chart_rows(
    monthly_values: list[float], year: int, *, baseline: float | None = None
) -> list[dict]:
    """Line chart rows with month labels and month-to-month change."""
    report_rows = []
    prev_balance_total = baseline if baseline is not None else 0.0
    for month_num in range(1, 13):
        total_bal = monthly_values[month_num - 1]
        change = total_bal - prev_balance_total
        prev_balance_total = total_bal
        report_rows.append(
            {
                "month_label": calendar.month_name[month_num],
                "month_num": month_num,
                "total_balance": total_bal,
                "change": change,
            }
        )
    return report_rows


def portfolio_value_from_holdings(holdings, price_by_key: dict, price_key: str) -> float:
    total = 0.0
    for h in holdings:
        key = h[price_key]
        pinfo = price_by_key.get(key)
        if pinfo and pinfo.get("price") is not None:
            total += float(h["quantity"]) * float(pinfo["price"])
        else:
            total += float(h["total_cost"])
    return total


def month_end_price_date_iso(year: int, month_num: int, today_d) -> str:
    """Calendar date used for month-end price (today for the in-progress current month)."""
    if year == today_d.year and month_num == today_d.month:
        return today_d.isoformat()
    if month_num == 12:
        next_month = datetime(year + 1, 1, 1)
    else:
        next_month = datetime(year, month_num + 1, 1)
    return (next_month - timedelta(days=1)).date().isoformat()


def _coingecko_history_date_param(iso_date: str) -> str:
    y, m, d = iso_date.split("-")
    return f"{int(d):02d}-{int(m):02d}-{y}"


_coingecko_history_last_at = 0.0
COINGECKO_HISTORY_MIN_INTERVAL = 0.35


def _throttle_coingecko_history():
    global _coingecko_history_last_at
    elapsed = time.time() - _coingecko_history_last_at
    if elapsed < COINGECKO_HISTORY_MIN_INTERVAL:
        time.sleep(COINGECKO_HISTORY_MIN_INTERVAL - elapsed)
    _coingecko_history_last_at = time.time()


def portfolio_baseline_before_year(
    conn,
    transactions,
    year: int,
    today_d,
    compute_holdings_fn,
    price_key: str,
    asset_kind: str,
) -> float:
    jan1 = f"{year:04d}-01-01"
    subset = [t for t in transactions if str(t["transacted_at"])[:10] < jan1]
    holdings = compute_holdings_fn(subset)
    if not holdings:
        return 0.0
    return portfolio_value_from_holdings(
        holdings,
        prices_for_holdings_at_month_end(
            conn, holdings, year - 1, 12, today_d, price_key, asset_kind
        ),
        price_key,
    )


def monthly_crypto_portfolio_values_for_year(conn, transactions, year: int, today_d) -> list[float]:
    balances = []
    for month_num in range(1, 13):
        snap_cutoff = month_snap_cutoff(year, month_num, today_d)
        subset = [t for t in transactions if str(t["transacted_at"])[:10] < snap_cutoff[:10]]
        holdings = compute_crypto_holdings(subset)
        price_map = prices_for_holdings_at_month_end(
            conn, holdings, year, month_num, today_d, "coin_id", "crypto"
        )
        balances.append(portfolio_value_from_holdings(holdings, price_map, "coin_id"))
    return balances


def monthly_stock_portfolio_values_for_year(conn, transactions, year: int, today_d) -> list[float]:
    balances = []
    for month_num in range(1, 13):
        snap_cutoff = month_snap_cutoff(year, month_num, today_d)
        subset = [t for t in transactions if str(t["transacted_at"])[:10] < snap_cutoff[:10]]
        holdings = compute_stock_holdings(subset)
        price_map = prices_for_holdings_at_month_end(
            conn, holdings, year, month_num, today_d, "symbol", "stock"
        )
        balances.append(portfolio_value_from_holdings(holdings, price_map, "symbol"))
    return balances


def normalize_reports_section(value):
    section = (value or "").strip().lower()
    return section if section in REPORTS_SECTIONS else "bank"


def normalize_report_account(value, conn, user_id):
    raw = (value or "").strip().lower()
    if not raw or raw == "all":
        return None
    try:
        account_id = int(raw)
    except ValueError:
        return None
    row = conn.execute(
        "SELECT id FROM accounts WHERE id = ? AND user_id = ?",
        (account_id, int(user_id)),
    ).fetchone()
    return int(row["id"]) if row else None


def balance_line_chart_spec(rows: list, *, width: float = 720, height: float = 300) -> dict:
    """SVG line chart: rows need total_balance, month_label, change."""
    ml, mr, mt, mb = 58.0, 18.0, 24.0, 46.0
    pw = width - ml - mr
    ph = height - mt - mb
    vals = [float(r["total_balance"]) for r in rows]
    month_label_y = height - 12.0
    if not vals:
        return {
            "w": width,
            "h": height,
            "points": "",
            "dots": [],
            "plot_ml": ml,
            "plot_mt": mt,
            "plot_pw": pw,
            "plot_ph": ph,
            "y_ticks": [],
            "month_label_y": month_label_y,
        }
    y_lo = min(vals)
    y_hi = max(vals)
    span = y_hi - y_lo
    if span <= 0:
        eps = max(abs(y_lo) * 0.02, 1.0) if y_lo != 0 else 1.0
        y_min = y_lo - eps
        y_max = y_hi + eps
    else:
        pad = span * 0.05
        y_min = y_lo - pad
        y_max = y_hi + pad
    y_rng = y_max - y_min or 1.0
    y_mid = y_min + y_rng / 2
    n = len(vals)
    dots = []
    pts = []
    for i, r in enumerate(rows):
        v = float(r["total_balance"])
        ch = float(r.get("change", 0.0))
        x = ml + (i / (n - 1)) * pw if n > 1 else ml + pw / 2
        yi = mt + ph - ((v - y_min) / y_rng) * ph
        pts.append(f"{x:.1f},{yi:.1f}")
        label = r["month_label"]
        dots.append(
            {
                "cx": round(x, 1),
                "cy": round(yi, 1),
                "lx": round(x, 1),
                "short": label[:3],
                "title": f"{label}: {v:.2f} (Δ {ch:+.2f})",
            }
        )
    y_ticks = [
        {"x": 4, "y": mt + 8, "text": f"{y_max:.2f}"},
        {"x": 4, "y": mt + ph / 2, "text": f"{y_mid:.2f}"},
        {"x": 4, "y": mt + ph - 4, "text": f"{y_min:.2f}"},
    ]
    return {
        "w": width,
        "h": height,
        "points": " ".join(pts),
        "dots": dots,
        "plot_ml": ml,
        "plot_mt": mt,
        "plot_pw": pw,
        "plot_ph": ph,
        "y_ticks": y_ticks,
        "month_label_y": month_label_y,
    }


def _avg_over_nonzero_month_cells(months: list[float]):
    """Mean of monthly values counting only months where the cell is non-zero."""
    active = sum(1 for v in months if v != 0.0)
    if not active:
        return None
    return sum(months) / active


def _empty_expense_pivot() -> dict:
    """Placeholder pivot used when the Reports panel is not being rendered."""
    return {
        "rows": [],
        "month_headers": [f"{m:02d}.{calendar.month_abbr[m]}" for m in range(1, 13)],
        "month_totals": [0.0] * 12,
        "grand_total": 0.0,
        "active_month_count": 0,
        "avg_monthly_total": None,
    }


def expense_pivot_for_report_year(conn, year: int, user_id: int) -> dict:
    """Category × month sums of expense amounts for a calendar year (values as stored in DB)."""
    uid = int(user_id)
    y0 = f"{year:04d}-01-01"
    y1 = f"{year + 1:04d}-01-01"
    raw = conn.execute(
        """
        SELECT c.name AS category_name,
               CAST(strftime('%m', e.spent_at) AS INTEGER) AS m,
               SUM(e.amount) AS total
        FROM expenses e
        JOIN categories c ON c.id = e.category_id AND c.user_id = e.user_id
        WHERE e.user_id = ? AND date(e.spent_at) >= date(?) AND date(e.spent_at) < date(?)
        GROUP BY c.name, m
        """,
        (uid, y0, y1),
    ).fetchall()
    pivot: dict[str, list[float]] = {}
    for row in raw:
        cat = row["category_name"]
        m = int(row["m"])
        if not 1 <= m <= 12:
            continue
        if cat not in pivot:
            pivot[cat] = [0.0] * 12
        pivot[cat][m - 1] = float(row["total"])
    categories_sorted = sorted(pivot.keys())
    rows_out = []
    for cat in categories_sorted:
        months = pivot[cat]
        rows_out.append(
            {
                "name": cat,
                "months": months,
                "total": sum(months),
                "avg": _avg_over_nonzero_month_cells(months),
            }
        )
    month_totals = [0.0] * 12
    for cat in categories_sorted:
        for i in range(12):
            month_totals[i] += pivot[cat][i]
    grand_total = sum(month_totals)
    month_headers = [f"{m:02d}.{calendar.month_abbr[m]}" for m in range(1, 13)]
    active_month_count = sum(1 for t in month_totals if t != 0.0)
    avg_monthly_total = (
        (grand_total / active_month_count) if active_month_count else None
    )
    return {
        "rows": rows_out,
        "month_headers": month_headers,
        "month_totals": month_totals,
        "grand_total": grand_total,
        "active_month_count": active_month_count,
        "avg_monthly_total": avg_monthly_total,
    }


def _column_names(conn, table):
    return {row["name"] for row in conn.execute(f"PRAGMA table_info({table})")}


def migrate_schema(conn):
    columns = _column_names(conn, "expenses")
    if columns and "notes" not in columns and "item" in columns:
        try:
            conn.execute("ALTER TABLE expenses RENAME COLUMN item TO notes")
        except sqlite3.OperationalError:
            conn.execute("ALTER TABLE expenses ADD COLUMN notes TEXT NOT NULL DEFAULT ''")
            conn.execute("UPDATE expenses SET notes = item WHERE notes = '' OR notes IS NULL")

    columns = _column_names(conn, "expenses")
    if columns and "account_id" not in columns:
        conn.execute("ALTER TABLE expenses ADD COLUMN account_id INTEGER REFERENCES accounts(id)")
        expense_account_id = conn.execute(
            "SELECT id FROM accounts WHERE name = 'Expense Cash' LIMIT 1"
        ).fetchone()
        if expense_account_id is None:
            conn.execute(
                "INSERT OR IGNORE INTO accounts(name, opening_balance) VALUES ('Expense Cash', 0)"
            )
            expense_account_id = conn.execute(
                "SELECT id FROM accounts WHERE name = 'Expense Cash' LIMIT 1"
            ).fetchone()
        fallback_account_id = expense_account_id["id"]
        conn.execute(
            "UPDATE expenses SET account_id = ? WHERE account_id IS NULL",
            (fallback_account_id,),
        )

    columns = _column_names(conn, "expenses")
    if columns and "spent_at" not in columns:
        conn.execute("ALTER TABLE expenses ADD COLUMN spent_at TIMESTAMP")
        conn.execute("UPDATE expenses SET spent_at = created_at WHERE spent_at IS NULL")

    columns = _column_names(conn, "income_entries")
    if columns and "notes" not in columns and "source" in columns:
        try:
            conn.execute("ALTER TABLE income_entries RENAME COLUMN source TO notes")
        except sqlite3.OperationalError:
            conn.execute("ALTER TABLE income_entries ADD COLUMN notes TEXT NOT NULL DEFAULT ''")
            conn.execute("UPDATE income_entries SET notes = source WHERE notes = '' OR notes IS NULL")

    columns = _column_names(conn, "income_entries")
    if columns and "category_id" not in columns:
        conn.execute("ALTER TABLE income_entries ADD COLUMN category_id INTEGER REFERENCES income_categories(id)")
        income_category_id = conn.execute(
            "SELECT id FROM income_categories WHERE name = 'General' LIMIT 1"
        ).fetchone()
        if income_category_id is None:
            conn.execute(
                "INSERT OR IGNORE INTO income_categories(name) VALUES ('General')"
            )
            income_category_id = conn.execute(
                "SELECT id FROM income_categories WHERE name = 'General' LIMIT 1"
            ).fetchone()
        fallback_category_id = income_category_id["id"]
        conn.execute(
            "UPDATE income_entries SET category_id = ? WHERE category_id IS NULL",
            (fallback_category_id,),
        )

    columns = _column_names(conn, "income_entries")
    if columns and "received_at" not in columns:
        conn.execute("ALTER TABLE income_entries ADD COLUMN received_at TIMESTAMP")
        conn.execute("UPDATE income_entries SET received_at = created_at WHERE received_at IS NULL")

    migrate_account_transfers(conn)
    migrate_expenses_signed_amounts(conn)
    migrate_txn_dates_to_day(conn)
    migrate_recurring_entries(conn)
    migrate_users_multitenancy(conn)
    migrate_crypto_transactions(conn)
    migrate_stock_transactions(conn)
    migrate_crypto_month_prices(conn)
    migrate_stock_month_prices(conn)
    integrations.migrate_user_integrations(conn)
    telegram_bot.migrate_telegram(conn)


def migrate_users_multitenancy(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE COLLATE NOCASE,
            password_hash TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    cat_cols = _column_names(conn, "categories") or []
    if "user_id" in cat_cols:
        return

    uid_row = conn.execute("SELECT id FROM users ORDER BY id LIMIT 1").fetchone()
    if uid_row:
        uid = int(uid_row["id"])
    else:
        legacy_pw = os.environ.get("VB_LEGACY_ADMIN_PASSWORD", "changeme")
        conn.execute(
            "INSERT INTO users (username, password_hash) VALUES (?, ?)",
            ("admin", generate_password_hash(legacy_pw)),
        )
        uid = int(conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"])

    conn.execute(
        """
        CREATE TABLE categories_new (
            id INTEGER PRIMARY KEY,
            user_id INTEGER NOT NULL REFERENCES users(id),
            name TEXT NOT NULL,
            UNIQUE(user_id, name)
        )
        """
    )
    conn.execute(
        "INSERT INTO categories_new (id, user_id, name) SELECT id, ?, name FROM categories",
        (uid,),
    )
    conn.execute("DROP TABLE categories")
    conn.execute("ALTER TABLE categories_new RENAME TO categories")

    conn.execute(
        """
        CREATE TABLE income_categories_new (
            id INTEGER PRIMARY KEY,
            user_id INTEGER NOT NULL REFERENCES users(id),
            name TEXT NOT NULL,
            UNIQUE(user_id, name)
        )
        """
    )
    conn.execute(
        "INSERT INTO income_categories_new (id, user_id, name) SELECT id, ?, name FROM income_categories",
        (uid,),
    )
    conn.execute("DROP TABLE income_categories")
    conn.execute("ALTER TABLE income_categories_new RENAME TO income_categories")

    conn.execute(
        """
        CREATE TABLE accounts_new (
            id INTEGER PRIMARY KEY,
            user_id INTEGER NOT NULL REFERENCES users(id),
            name TEXT NOT NULL,
            opening_balance REAL NOT NULL DEFAULT 0,
            UNIQUE(user_id, name)
        )
        """
    )
    conn.execute(
        """
        INSERT INTO accounts_new (id, user_id, name, opening_balance)
        SELECT id, ?, name, opening_balance FROM accounts
        """,
        (uid,),
    )
    conn.execute("DROP TABLE accounts")
    conn.execute("ALTER TABLE accounts_new RENAME TO accounts")

    conn.execute(f"ALTER TABLE expenses ADD COLUMN user_id INTEGER NOT NULL DEFAULT {uid}")
    conn.execute(f"ALTER TABLE income_entries ADD COLUMN user_id INTEGER NOT NULL DEFAULT {uid}")
    conn.execute(f"ALTER TABLE account_transfers ADD COLUMN user_id INTEGER NOT NULL DEFAULT {uid}")
    conn.execute(f"ALTER TABLE recurring_entries ADD COLUMN user_id INTEGER NOT NULL DEFAULT {uid}")


def seed_user_defaults(conn, user_id):
    uid = int(user_id)
    conn.execute(
        "INSERT OR IGNORE INTO categories (user_id, name) VALUES (?, ?)",
        (uid, "General"),
    )
    conn.execute(
        "INSERT OR IGNORE INTO categories (user_id, name) VALUES (?, ?)",
        (uid, "Other"),
    )
    conn.execute(
        "INSERT OR IGNORE INTO income_categories (user_id, name) VALUES (?, ?)",
        (uid, "General"),
    )
    conn.execute(
        "INSERT OR IGNORE INTO income_categories (user_id, name) VALUES (?, ?)",
        (uid, "Salary"),
    )
    conn.execute(
        "INSERT OR IGNORE INTO income_categories (user_id, name) VALUES (?, ?)",
        (uid, "Other"),
    )
    conn.execute(
        "INSERT OR IGNORE INTO accounts (user_id, name, opening_balance) VALUES (?, ?, ?)",
        (uid, "Main", 0.0),
    )


def migrate_crypto_transactions(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crypto_transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL REFERENCES users(id),
            coin_id TEXT NOT NULL,
            coin_symbol TEXT NOT NULL,
            coin_name TEXT NOT NULL,
            tx_type TEXT NOT NULL CHECK (tx_type IN ('buy', 'sell')),
            quantity REAL NOT NULL CHECK (quantity > 0),
            price_per_unit REAL NOT NULL CHECK (price_per_unit >= 0),
            fee REAL NOT NULL DEFAULT 0 CHECK (fee >= 0),
            exchange TEXT NOT NULL DEFAULT '',
            transacted_at TEXT NOT NULL,
            notes TEXT NOT NULL DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
    )


_price_cache: dict = {"prices": {}, "fetched_at": 0.0}
_price_cache_lock = threading.Lock()
PRICE_CACHE_TTL = 300


def fetch_coingecko_prices(coin_ids, force=False):
    if not coin_ids:
        return {}
    now = time.time()
    cached = _price_cache
    with _price_cache_lock:
        if not force and cached["prices"] and (now - cached["fetched_at"]) < PRICE_CACHE_TTL:
            if all(cid in cached["prices"] for cid in coin_ids):
                return {cid: cached["prices"][cid] for cid in coin_ids}
    ids_str = ",".join(sorted(set(coin_ids)))
    url = f"https://api.coingecko.com/api/v3/simple/price?ids={ids_str}&vs_currencies=eur&include_24hr_change=true"
    try:
        req = Request(url, headers={"Accept": "application/json", "User-Agent": "VibeBudgeting/1.0"})
        with urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode())
        prices = {}
        for cid in coin_ids:
            if cid in data and "eur" in data[cid]:
                prices[cid] = {
                    "price": data[cid]["eur"],
                    "change_24h": data[cid].get("eur_24h_change"),
                }
        with _price_cache_lock:
            cached["prices"].update(prices)
            cached["fetched_at"] = time.time()
        return prices
    except (URLError, HTTPError, json.JSONDecodeError, OSError):
        with _price_cache_lock:
            return {cid: cached["prices"][cid] for cid in coin_ids if cid in cached["prices"]}


def migrate_crypto_month_prices(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crypto_month_prices (
            coin_id TEXT NOT NULL,
            year INTEGER NOT NULL,
            month INTEGER NOT NULL,
            price_eur REAL NOT NULL,
            price_date TEXT NOT NULL,
            source TEXT NOT NULL DEFAULT 'coingecko',
            fetched_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (coin_id, year, month)
        )
        """
    )


def migrate_stock_month_prices(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS stock_month_prices (
            symbol TEXT NOT NULL,
            year INTEGER NOT NULL,
            month INTEGER NOT NULL,
            price_usd REAL NOT NULL,
            price_date TEXT NOT NULL,
            source TEXT NOT NULL DEFAULT 'yfinance',
            fetched_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (symbol, year, month)
        )
        """
    )


def fetch_coingecko_history_eur(coin_id: str, price_date_iso: str) -> float | None:
    """CoinGecko daily snapshot for a calendar date (dd-mm-yyyy query param)."""
    date_param = _coingecko_history_date_param(price_date_iso)
    url = f"https://api.coingecko.com/api/v3/coins/{coin_id}/history?date={date_param}"
    _throttle_coingecko_history()
    try:
        req = Request(url, headers={"Accept": "application/json", "User-Agent": "VibeBudgeting/1.0"})
        with urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode())
        market = data.get("market_data") or {}
        current = market.get("current_price") or {}
        eur = current.get("eur")
        if eur is not None and float(eur) > 0:
            return float(eur)
    except (URLError, HTTPError, json.JSONDecodeError, OSError, ValueError, TypeError):
        return None
    return None


def fetch_stock_month_close_usd(symbol: str, price_date_iso: str) -> float | None:
    """Last available close in the month window up to price_date_iso, in USD."""
    lookup = _quote_lookup_symbol(symbol)
    try:
        import yfinance as yf

        end_d = datetime.strptime(price_date_iso, "%Y-%m-%d").date()
        start_d = end_d.replace(day=1)
        hist = yf.Ticker(lookup).history(
            start=start_d.isoformat(),
            end=(end_d + timedelta(days=1)).isoformat(),
            auto_adjust=True,
        )
        if hist is None or hist.empty:
            return None
        close = float(hist["Close"].iloc[-1])
        if close <= 0:
            return None
        return _listing_price_to_usd(lookup, close)
    except Exception:
        return None


def _is_current_report_month(year: int, month_num: int, today_d) -> bool:
    return year == today_d.year and month_num == today_d.month


def ensure_crypto_month_price(conn, coin_id: str, year: int, month_num: int, today_d):
    if _is_current_report_month(year, month_num, today_d):
        live = fetch_coingecko_prices([coin_id], force=False)
        return live.get(coin_id)

    row = conn.execute(
        """
        SELECT price_eur FROM crypto_month_prices
        WHERE coin_id = ? AND year = ? AND month = ?
        """,
        (coin_id, year, month_num),
    ).fetchone()
    if row:
        return {"price": float(row["price_eur"]), "source": "cache"}

    price_date = month_end_price_date_iso(year, month_num, today_d)
    price = fetch_coingecko_history_eur(coin_id, price_date)
    source = "coingecko"
    if price is None:
        live = fetch_coingecko_prices([coin_id], force=False)
        pinfo = live.get(coin_id)
        if pinfo:
            price = float(pinfo["price"])
            source = "coingecko_live_fallback"
    if price is None:
        return None

    conn.execute(
        """
        INSERT OR REPLACE INTO crypto_month_prices
            (coin_id, year, month, price_eur, price_date, source)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (coin_id, year, month_num, price, price_date, source),
    )
    return {"price": price, "source": source}


def ensure_stock_month_price(conn, symbol: str, year: int, month_num: int, today_d):
    if _is_current_report_month(year, month_num, today_d):
        live = fetch_finnhub_quotes([symbol], force=False)
        return live.get(symbol)

    row = conn.execute(
        """
        SELECT price_usd FROM stock_month_prices
        WHERE symbol = ? AND year = ? AND month = ?
        """,
        (symbol, year, month_num),
    ).fetchone()
    if row:
        return {"price": float(row["price_usd"]), "source": "cache"}

    price_date = month_end_price_date_iso(year, month_num, today_d)
    price = fetch_stock_month_close_usd(symbol, price_date)
    source = "yfinance"
    if price is None:
        live = fetch_finnhub_quotes([symbol], force=False)
        pinfo = live.get(symbol)
        if pinfo:
            price = float(pinfo["price"])
            source = "finnhub_live_fallback"
    if price is None:
        return None

    conn.execute(
        """
        INSERT OR REPLACE INTO stock_month_prices
            (symbol, year, month, price_usd, price_date, source)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (symbol, year, month_num, price, price_date, source),
    )
    return {"price": price, "source": source}


def prices_for_holdings_at_month_end(
    conn, holdings, year: int, month_num: int, today_d, price_key: str, asset_kind: str
) -> dict:
    price_map = {}
    for h in holdings:
        key = h[price_key]
        if asset_kind == "crypto":
            pinfo = ensure_crypto_month_price(conn, key, year, month_num, today_d)
        else:
            pinfo = ensure_stock_month_price(conn, key, year, month_num, today_d)
        if pinfo:
            price_map[key] = pinfo
    return price_map


def compute_crypto_holdings(transactions):
    holdings: dict = {}
    for tx in transactions:
        cid = tx["coin_id"]
        if cid not in holdings:
            holdings[cid] = {
                "coin_id": cid,
                "coin_symbol": tx["coin_symbol"],
                "coin_name": tx["coin_name"],
                "quantity": 0.0,
                "total_cost": 0.0,
            }
        h = holdings[cid]
        qty = float(tx["quantity"])
        price = float(tx["price_per_unit"])
        fee = float(tx["fee"])
        if tx["tx_type"] == "buy":
            h["total_cost"] += qty * price + fee
            h["quantity"] += qty
        else:
            if h["quantity"] > 0:
                avg_cost = h["total_cost"] / h["quantity"]
                h["total_cost"] -= qty * avg_cost
            h["quantity"] -= qty
    result = []
    for _cid, h in sorted(holdings.items(), key=lambda x: x[1]["coin_name"].lower()):
        if abs(h["quantity"]) > 1e-9:
            h["avg_buy_price"] = h["total_cost"] / h["quantity"] if h["quantity"] > 0 else 0
            result.append(h)
    return result


def normalize_investments_section(value):
    section = (value or "").strip().lower()
    return section if section in INVESTMENTS_SECTIONS else "crypto"


def migrate_stock_transactions(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS stock_transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL REFERENCES users(id),
            symbol TEXT NOT NULL,
            ticker TEXT NOT NULL,
            instrument_name TEXT NOT NULL,
            tx_type TEXT NOT NULL CHECK (tx_type IN ('buy', 'sell')),
            quantity REAL NOT NULL CHECK (quantity > 0),
            price_per_unit REAL NOT NULL CHECK (price_per_unit >= 0),
            fee REAL NOT NULL DEFAULT 0 CHECK (fee >= 0),
            broker TEXT NOT NULL DEFAULT '',
            transacted_at TEXT NOT NULL,
            notes TEXT NOT NULL DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
    )


_stock_price_cache: dict = {"prices": {}, "fetched_at": 0.0}
_stock_price_cache_lock = threading.Lock()
_fx_usd_cache: dict = {"rates": {}, "fetched_at": 0.0}
_fx_usd_cache_lock = threading.Lock()
_EUR_LISTING_SUFFIXES = (
    ".DE",
    ".AS",
    ".PA",
    ".MI",
    ".HE",
    ".SW",
    ".BR",
    ".VI",
    ".OL",
    ".ST",
    ".CO",
    ".IR",
    ".LS",
    ".MC",
    ".BE",
    ".WA",
)
_GBP_LISTING_SUFFIXES = (".L",)
# Freedom24 uses .EU on USD-denominated UCITS listings (not a Yahoo/Finnhub symbol).
_FREEDOM_EU_SUFFIX = ".EU"


def _quote_lookup_symbol(symbol):
    """Map broker symbols to a symbol Yahoo/Finnhub can quote."""
    sym = (symbol or "").strip().upper()
    if sym.endswith(_FREEDOM_EU_SUFFIX):
        return f"{sym[: -len(_FREEDOM_EU_SUFFIX)]}.DE"
    return sym


def _finnhub_request(path):
    token = FINNHUB_API_KEY
    if not token:
        return None
    sep = "&" if "?" in path else "?"
    url = f"https://finnhub.io/api/v1{path}{sep}token={token}"
    try:
        req = Request(url, headers={"Accept": "application/json", "User-Agent": "VibeBudgeting/1.0"})
        with urlopen(req, timeout=10) as resp:
            return json.loads(resp.read().decode())
    except (URLError, HTTPError, json.JSONDecodeError, OSError):
        return None


def _yfinance_last_price(symbol):
    try:
        import yfinance as yf

        ticker = yf.Ticker(symbol)
        price = getattr(ticker, "fast_info", {}).get("lastPrice")
        if price is not None and float(price) > 0:
            return float(price)
        hist = ticker.history(period="5d")
        if hist is not None and not hist.empty:
            close = float(hist["Close"].iloc[-1])
            if close > 0:
                return close
    except Exception:
        return None
    return None


def _finnhub_fx_to_usd(pair):
    """OANDA:EUR_USD / OANDA:GBP_USD → USD per 1 unit of base currency."""
    now = time.time()
    cached = _fx_usd_cache
    with _fx_usd_cache_lock:
        if not cached["rates"] or (now - cached["fetched_at"]) >= PRICE_CACHE_TTL:
            cached["rates"] = {}
            cached["fetched_at"] = now
        if pair in cached["rates"]:
            return cached["rates"][pair]
    data = _finnhub_request(f"/quote?symbol={pair}")
    if data and data.get("c") is not None and float(data["c"]) > 0:
        rate = float(data["c"])
        with _fx_usd_cache_lock:
            cached["rates"][pair] = rate
        return rate
    yf_symbol = "EURUSD=X" if "EUR" in pair else "GBPUSD=X"
    yf_rate = _yfinance_last_price(yf_symbol)
    if yf_rate:
        with _fx_usd_cache_lock:
            cached["rates"][pair] = yf_rate
        return yf_rate
    return None


def _listing_price_to_usd(symbol, price):
    sym = symbol.upper()
    if any(sym.endswith(suf) for suf in _GBP_LISTING_SUFFIXES):
        rate = _finnhub_fx_to_usd("OANDA:GBP_USD")
        if rate:
            return price * rate
    if any(sym.endswith(suf) for suf in _EUR_LISTING_SUFFIXES):
        rate = _finnhub_fx_to_usd("OANDA:EUR_USD")
        if rate:
            return price * rate
    return price


def fetch_finnhub_quotes(symbols, force=False):
    if not symbols:
        return {}
    now = time.time()
    cached = _stock_price_cache
    with _stock_price_cache_lock:
        if not force and cached["prices"] and (now - cached["fetched_at"]) < PRICE_CACHE_TTL:
            if all(sym in cached["prices"] for sym in symbols):
                return {sym: cached["prices"][sym] for sym in symbols}
    if force:
        with _fx_usd_cache_lock:
            _fx_usd_cache["rates"] = {}
            _fx_usd_cache["fetched_at"] = 0.0
    prices = {}
    for sym in sorted(set(symbols)):
        raw = None
        change_24h = None
        lookup = _quote_lookup_symbol(sym)
        data = _finnhub_request(f"/quote?symbol={lookup}") if FINNHUB_API_KEY else None
        if data and data.get("c") is not None and float(data["c"]) > 0:
            raw = float(data["c"])
            change_24h = float(data["dp"]) if data.get("dp") is not None else None
        if raw is None:
            yf_raw = _yfinance_last_price(lookup)
            if yf_raw is not None and yf_raw > 0:
                raw = yf_raw
        if raw is not None and raw > 0:
            prices[sym] = {
                "price": _listing_price_to_usd(lookup, raw),
                "change_24h": change_24h,
                "source": "finnhub" if data and data.get("c") else "yfinance",
            }
    with _stock_price_cache_lock:
        cached["prices"].update(prices)
        cached["fetched_at"] = time.time()
    return prices


def compute_stock_holdings(transactions):
    holdings: dict = {}
    for tx in transactions:
        sym = tx["symbol"]
        if sym not in holdings:
            holdings[sym] = {
                "symbol": sym,
                "ticker": tx["ticker"],
                "instrument_name": tx["instrument_name"],
                "quantity": 0.0,
                "total_cost": 0.0,
            }
        h = holdings[sym]
        qty = float(tx["quantity"])
        price = float(tx["price_per_unit"])
        fee = float(tx["fee"])
        if tx["tx_type"] == "buy":
            h["total_cost"] += qty * price + fee
            h["quantity"] += qty
        else:
            if h["quantity"] > 0:
                avg_cost = h["total_cost"] / h["quantity"]
                h["total_cost"] -= qty * avg_cost
            h["quantity"] -= qty
    result = []
    for _sym, h in sorted(holdings.items(), key=lambda x: x[1]["instrument_name"].lower()):
        if abs(h["quantity"]) > 1e-9:
            h["avg_buy_price"] = h["total_cost"] / h["quantity"] if h["quantity"] > 0 else 0
            result.append(h)
    return result


def migrate_recurring_entries(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS recurring_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entry_type TEXT NOT NULL CHECK (entry_type IN ('expense', 'income')),
            amount REAL NOT NULL CHECK (amount > 0),
            category_id INTEGER NOT NULL,
            account_id INTEGER NOT NULL REFERENCES accounts(id),
            day_of_month INTEGER NOT NULL CHECK (day_of_month >= 1 AND day_of_month <= 31),
            notes TEXT NOT NULL DEFAULT '',
            enabled INTEGER NOT NULL DEFAULT 1 CHECK (enabled IN (0, 1)),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS recurring_applied (
            recurring_id INTEGER NOT NULL,
            ym TEXT NOT NULL,
            PRIMARY KEY (recurring_id, ym),
            FOREIGN KEY (recurring_id) REFERENCES recurring_entries(id) ON DELETE CASCADE
        )
        """
    )


def migrate_account_transfers(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS account_transfers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            from_account_id INTEGER NOT NULL,
            to_account_id INTEGER NOT NULL,
            amount REAL NOT NULL CHECK (amount > 0),
            transferred_at TEXT NOT NULL,
            notes TEXT NOT NULL DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (from_account_id) REFERENCES accounts(id),
            FOREIGN KEY (to_account_id) REFERENCES accounts(id),
            CHECK (from_account_id != to_account_id)
        )
        """
    )


def migrate_txn_dates_to_day(conn):
    """Store movement dates as calendar days only (YYYY-MM-DD)."""
    conn.execute(
        """
        UPDATE expenses
        SET spent_at = date(spent_at)
        WHERE spent_at IS NOT NULL AND trim(spent_at) != '' AND date(spent_at) IS NOT NULL
        """
    )
    conn.execute(
        """
        UPDATE income_entries
        SET received_at = date(received_at)
        WHERE received_at IS NOT NULL AND trim(received_at) != '' AND date(received_at) IS NOT NULL
        """
    )


def migrate_expenses_signed_amounts(conn):
    row = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='expenses'"
    ).fetchone()
    if not row or not row["sql"]:
        return
    create_sql = row["sql"]
    if not re.search(r"CHECK\s*\(\s*amount\s*>=\s*0\s*\)", create_sql, re.I):
        return
    cols = _column_names(conn, "expenses") or []
    has_uid = "user_id" in cols
    if has_uid:
        conn.executescript(
            """
            CREATE TABLE expenses_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                notes TEXT NOT NULL DEFAULT '',
                amount REAL NOT NULL,
                category_id INTEGER NOT NULL,
                account_id INTEGER NOT NULL,
                spent_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (category_id) REFERENCES categories(id),
                FOREIGN KEY (account_id) REFERENCES accounts(id)
            );
            INSERT INTO expenses_new (id, user_id, notes, amount, category_id, account_id, spent_at, created_at)
            SELECT id, user_id, notes, -ABS(amount), category_id, account_id, spent_at, created_at FROM expenses;
            DROP TABLE expenses;
            ALTER TABLE expenses_new RENAME TO expenses;
            """
        )
    else:
        conn.executescript(
            """
            CREATE TABLE expenses_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                notes TEXT NOT NULL DEFAULT '',
                amount REAL NOT NULL,
                category_id INTEGER NOT NULL,
                account_id INTEGER NOT NULL,
                spent_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (category_id) REFERENCES categories(id),
                FOREIGN KEY (account_id) REFERENCES accounts(id)
            );
            INSERT INTO expenses_new (id, notes, amount, category_id, account_id, spent_at, created_at)
            SELECT id, notes, -ABS(amount), category_id, account_id, spent_at, created_at FROM expenses;
            DROP TABLE expenses;
            ALTER TABLE expenses_new RENAME TO expenses;
            """
        )


def init_db():
    print(f"[vibe-budgeting] SQLite database path: {DB_PATH!r}", file=sys.stderr)
    conn = get_connection()
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        );

        CREATE TABLE IF NOT EXISTS income_categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        );

        CREATE TABLE IF NOT EXISTS accounts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            opening_balance REAL NOT NULL DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            notes TEXT NOT NULL DEFAULT '',
            amount REAL NOT NULL,
            category_id INTEGER NOT NULL,
            account_id INTEGER NOT NULL,
            spent_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (category_id) REFERENCES categories(id),
            FOREIGN KEY (account_id) REFERENCES accounts(id)
        );

        CREATE TABLE IF NOT EXISTS income_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            notes TEXT NOT NULL DEFAULT '',
            amount REAL NOT NULL CHECK (amount >= 0),
            category_id INTEGER NOT NULL,
            account_id INTEGER NOT NULL,
            received_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (category_id) REFERENCES income_categories(id),
            FOREIGN KEY (account_id) REFERENCES accounts(id)
        );
        """
    )

    conn.execute("INSERT OR IGNORE INTO categories(name) VALUES ('General')")
    conn.execute("INSERT OR IGNORE INTO income_categories(name) VALUES ('General')")
    conn.execute("INSERT OR IGNORE INTO accounts(name, opening_balance) VALUES ('Main', 0)")
    migrate_schema(conn)
    conn.commit()
    conn.close()


def resolve_active_panel():
    panel = request.args.get("panel", "").strip()
    if panel in ALLOWED_PANELS:
        return panel

    next_panel = request.form.get("next_panel", "").strip()
    if next_panel in ALLOWED_PANELS:
        return next_panel

    referer = request.headers.get("Referer", "")
    if referer:
        path = urlparse(referer).path.rstrip("/") or "/"
        mapping = {
            "/categories/add": "settings",
            "/expenses/add": "expenses",
            "/income/add": "income",
            "/transfers/add": "transfer",
            "/import/excel": "settings",
        }
        if path in mapping:
            return mapping[path]

        for prefix, mapped in (
            ("/categories/", "settings"),
            ("/income-categories/", "settings"),
            ("/accounts/", "settings"),
            ("/expenses/", "expenses"),
            ("/income/", "income"),
            ("/transfers/", "transfer"),
            ("/recurring/", "recurring"),
            ("/crypto/", "investments"),
            ("/stocks/", "investments"),
        ):
            if path.startswith(prefix):
                return mapped

    return "home"


def redirect_home(panel=None, settings_section=None):
    target = panel if panel in ALLOWED_PANELS else resolve_active_panel()
    month = normalize_month(request.form.get("month") or request.args.get("month"))
    year_for_redirect = normalize_year(request.form.get("year") or request.args.get("year"))
    raw_section = (
        settings_section
        if settings_section is not None
        else (request.form.get("settings_section") or request.args.get("settings_section"))
    )
    sec = normalize_settings_section(raw_section)
    query = {"panel": target, "month": month}
    if target == "settings":
        query["settings_section"] = sec
    if target == "investments":
        raw_inv_sec = (
            settings_section
            if settings_section is not None
            else (request.form.get("investments_section") or request.args.get("investments_section"))
        )
        query["investments_section"] = normalize_investments_section(raw_inv_sec)
    if target == "yearly":
        query["year"] = year_for_redirect
    report_year_redirect = normalize_year(
        request.form.get("report_year") or request.args.get("report_year")
    )
    if target == "reports":
        query["report_year"] = report_year_redirect
        raw_reports_sec = request.form.get("reports_section") or request.args.get("reports_section")
        if raw_reports_sec:
            query["reports_section"] = normalize_reports_section(raw_reports_sec)
        raw_report_acct = request.form.get("report_account") or request.args.get("report_account")
        if raw_report_acct is not None and str(raw_report_acct).strip():
            query["report_account"] = str(raw_report_acct).strip()
    exp_pg = normalize_list_page(request.form.get("exp_page") or request.args.get("exp_page") or 1)
    inc_pg = normalize_list_page(request.form.get("inc_page") or request.args.get("inc_page") or 1)
    if target == "expenses" and exp_pg > 1:
        query["exp_page"] = exp_pg
    if target == "income" and inc_pg > 1:
        query["inc_page"] = inc_pg
    exp_fm = parse_optional_month(request.form.get("exp_month") or request.args.get("exp_month"))
    inc_fm = parse_optional_month(request.form.get("inc_month") or request.args.get("inc_month"))
    if target == "expenses" and exp_fm:
        query["exp_month"] = exp_fm
    if target == "income" and inc_fm:
        query["inc_month"] = inc_fm
    exp_cat_id = normalize_optional_category_id(
        request.form.get("exp_category") or request.args.get("exp_category")
    )
    if target == "expenses" and exp_cat_id is not None:
        query["exp_category"] = exp_cat_id
    inc_cat_id = normalize_optional_category_id(
        request.form.get("inc_category") or request.args.get("inc_category")
    )
    if target == "income" and inc_cat_id is not None:
        query["inc_category"] = inc_cat_id
    return redirect(url_for("index", **query))


def _normalize_header_key(value):
    if value is None:
        return ""
    return str(value).strip().lower()


def _sheet_as_dicts(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers_raw = [_normalize_header_key(cell) for cell in rows[0]]
    headers = []
    for raw in headers_raw:
        headers.append(raw if raw else "")
    out = []
    for row in rows[1:]:
        if row is None:
            continue
        cells = list(row)
        if not cells:
            continue
        if all(cell is None or str(cell).strip() == "" for cell in cells):
            continue
        row_dict = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row_dict[header] = cells[idx] if idx < len(cells) else None
        out.append(row_dict)
    return out


def _parse_excel_expense_amount(value, sheet, row_num):
    """Excel import: sign is kept. Negative = spending; positive = refund/credit (adds back to balance)."""
    if value is None or (isinstance(value, str) and not value.strip()):
        raise ValueError(f"{sheet} row {row_num}: missing amount")
    try:
        amount = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{sheet} row {row_num}: invalid amount") from exc
    return amount


def _parse_excel_income_amount(value, sheet, row_num):
    if value is None or (isinstance(value, str) and not value.strip()):
        raise ValueError(f"{sheet} row {row_num}: missing amount")
    try:
        amount = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{sheet} row {row_num}: invalid amount") from exc
    return abs(amount)


def _parse_excel_datetime(value, sheet, row_num, column_label):
    if value is None or (isinstance(value, str) and not value.strip()):
        raise ValueError(f"{sheet} row {row_num}: missing {column_label}")
    if isinstance(value, datetime):
        return value.replace(microsecond=0)
    text = str(value).strip()
    try:
        normalized = text.replace(" ", "T", 1)
        if len(normalized) == 10:
            normalized = f"{normalized}T00:00:00"
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        parsed = None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                parsed = datetime.strptime(text, fmt)
                break
            except ValueError:
                continue
        if parsed is None:
            raise ValueError(f"{sheet} row {row_num}: invalid {column_label}")
    return parsed.replace(microsecond=0)


def _parse_excel_timestamp(value, sheet, row_num, column_label):
    return _parse_excel_datetime(value, sheet, row_num, column_label).isoformat(timespec="seconds")


def _parse_excel_movement_date(value, sheet, row_num, column_label):
    return _parse_excel_datetime(value, sheet, row_num, column_label).date().isoformat()


def _optional_created_at(value, sheet, row_num):
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    return _parse_excel_timestamp(value, sheet, row_num, "created_at")


def _build_export_workbook(conn, user_id):
    uid = int(user_id)
    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = SHEET_META
    ws_meta.append(["key", "value"])
    ws_meta.append(["format_version", EXPORT_FORMAT_VERSION])
    ws_meta.append(["exported_at", datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")])

    ws_accounts = wb.create_sheet(SHEET_ACCOUNTS)
    ws_accounts.append(["name", "opening_balance"])
    for row in conn.execute(
        "SELECT name, opening_balance FROM accounts WHERE user_id = ? ORDER BY name",
        (uid,),
    ):
        ws_accounts.append([row["name"], row["opening_balance"]])

    ws_ec = wb.create_sheet(SHEET_EXPENSE_CATEGORIES)
    ws_ec.append(["name"])
    for row in conn.execute(
        "SELECT name FROM categories WHERE user_id = ? ORDER BY name",
        (uid,),
    ):
        ws_ec.append([row["name"]])

    ws_ic = wb.create_sheet(SHEET_INCOME_CATEGORIES)
    ws_ic.append(["name"])
    for row in conn.execute(
        "SELECT name FROM income_categories WHERE user_id = ? ORDER BY name",
        (uid,),
    ):
        ws_ic.append([row["name"]])

    ws_exp = wb.create_sheet(SHEET_EXPENSES)
    ws_exp.append(["notes", "amount", "category_name", "account_name", "spent_at", "created_at"])
    for row in conn.execute(
        """
        SELECT e.notes, e.amount, c.name AS category_name, a.name AS account_name, e.spent_at, e.created_at
        FROM expenses e
        JOIN categories c ON c.id = e.category_id
        JOIN accounts a ON a.id = e.account_id
        WHERE e.user_id = ?
        ORDER BY e.spent_at ASC, e.id ASC
        """,
        (uid,),
    ):
        ws_exp.append(
            [
                row["notes"],
                row["amount"],
                row["category_name"],
                row["account_name"],
                row["spent_at"],
                row["created_at"],
            ]
        )

    ws_inc = wb.create_sheet(SHEET_INCOME)
    ws_inc.append(["notes", "amount", "category_name", "account_name", "received_at", "created_at"])
    for row in conn.execute(
        """
        SELECT i.notes, i.amount, c.name AS category_name, a.name AS account_name, i.received_at, i.created_at
        FROM income_entries i
        JOIN income_categories c ON c.id = i.category_id
        JOIN accounts a ON a.id = i.account_id
        WHERE i.user_id = ?
        ORDER BY i.received_at ASC, i.id ASC
        """,
        (uid,),
    ):
        ws_inc.append(
            [
                row["notes"],
                row["amount"],
                row["category_name"],
                row["account_name"],
                row["received_at"],
                row["created_at"],
            ]
        )

    return wb


def _build_migration_template_workbook():
    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = SHEET_META
    ws_meta.append(["key", "value"])
    ws_meta.append(["format_version", EXPORT_FORMAT_VERSION])
    ws_meta.append(["kind", "migration_template"])

    ws_accounts = wb.create_sheet(SHEET_ACCOUNTS)
    ws_accounts.append(["name", "opening_balance"])
    ws_accounts.append(["Main", 0])

    ws_ec = wb.create_sheet(SHEET_EXPENSE_CATEGORIES)
    ws_ec.append(["name"])
    ws_ec.append(["General"])

    ws_ic = wb.create_sheet(SHEET_INCOME_CATEGORIES)
    ws_ic.append(["name"])
    ws_ic.append(["General"])

    ws_exp = wb.create_sheet(SHEET_EXPENSES)
    ws_exp.append(["notes", "amount", "category_name", "account_name", "spent_at", "created_at"])

    ws_inc = wb.create_sheet(SHEET_INCOME)
    ws_inc.append(["notes", "amount", "category_name", "account_name", "received_at", "created_at"])

    return wb


def _lookup_category_id(conn, name, user_id, expense=True):
    uid = int(user_id)
    table = "categories" if expense else "income_categories"
    row = conn.execute(
        f"SELECT id FROM {table} WHERE user_id = ? AND name = ?",
        (uid, name.strip()),
    ).fetchone()
    return row["id"] if row else None


def _lookup_account_id(conn, name, user_id):
    uid = int(user_id)
    row = conn.execute(
        "SELECT id FROM accounts WHERE user_id = ? AND name = ?",
        (uid, name.strip()),
    ).fetchone()
    return row["id"] if row else None


def _user_owns_category(conn, category_id, user_id, *, expense):
    table = "categories" if expense else "income_categories"
    return (
        conn.execute(
            f"SELECT 1 FROM {table} WHERE id = ? AND user_id = ?",
            (int(category_id), int(user_id)),
        ).fetchone()
        is not None
    )


def _user_owns_account(conn, account_id, user_id):
    return (
        conn.execute(
            "SELECT 1 FROM accounts WHERE id = ? AND user_id = ?",
            (int(account_id), int(user_id)),
        ).fetchone()
        is not None
    )


def _safe_next_url(raw):
    if not raw or not isinstance(raw, str):
        return None
    text = raw.strip()
    if not text.startswith("/") or text.startswith("//"):
        return None
    return text


def _collect_import_movements(expense_rows, income_rows):
    errors = []
    insert_expenses = []
    for idx, row in enumerate(expense_rows, start=2):
        notes_val = row.get("notes")
        notes = "" if notes_val is None else str(notes_val)
        try:
            amount = _parse_excel_expense_amount(row.get("amount"), SHEET_EXPENSES, idx)
        except ValueError as exc:
            errors.append(str(exc))
            continue
        cat_name = row.get("category_name")
        acc_name = row.get("account_name")
        if cat_name is None or not str(cat_name).strip():
            errors.append(f"{SHEET_EXPENSES} row {idx}: missing category_name")
            continue
        if acc_name is None or not str(acc_name).strip():
            errors.append(f"{SHEET_EXPENSES} row {idx}: missing account_name")
            continue
        cat_name = str(cat_name).strip()
        acc_name = str(acc_name).strip()
        try:
            spent_at = _parse_excel_movement_date(row.get("spent_at"), SHEET_EXPENSES, idx, "spent_at")
        except ValueError as exc:
            errors.append(str(exc))
            continue
        created_at = None
        try:
            if row.get("created_at") not in (None, ""):
                created_at = _optional_created_at(row.get("created_at"), SHEET_EXPENSES, idx)
        except ValueError as exc:
            errors.append(str(exc))
            continue
        insert_expenses.append((notes, amount, cat_name, acc_name, spent_at, created_at))

    insert_income = []
    for idx, row in enumerate(income_rows, start=2):
        notes_val = row.get("notes")
        notes = "" if notes_val is None else str(notes_val)
        try:
            amount = _parse_excel_income_amount(row.get("amount"), SHEET_INCOME, idx)
        except ValueError as exc:
            errors.append(str(exc))
            continue
        cat_name = row.get("category_name")
        acc_name = row.get("account_name")
        if cat_name is None or not str(cat_name).strip():
            errors.append(f"{SHEET_INCOME} row {idx}: missing category_name")
            continue
        if acc_name is None or not str(acc_name).strip():
            errors.append(f"{SHEET_INCOME} row {idx}: missing account_name")
            continue
        cat_name = str(cat_name).strip()
        acc_name = str(acc_name).strip()
        try:
            received_at = _parse_excel_movement_date(
                row.get("received_at"), SHEET_INCOME, idx, "received_at"
            )
        except ValueError as exc:
            errors.append(str(exc))
            continue
        created_at = None
        try:
            if row.get("created_at") not in (None, ""):
                created_at = _optional_created_at(row.get("created_at"), SHEET_INCOME, idx)
        except ValueError as exc:
            errors.append(str(exc))
            continue
        insert_income.append((notes, amount, cat_name, acc_name, received_at, created_at))

    return insert_expenses, insert_income, errors


def _run_import_workbook(wb, replace_movements, sync_opening_balances, user_id):
    errors = []
    required_sheets = {
        SHEET_ACCOUNTS,
        SHEET_EXPENSE_CATEGORIES,
        SHEET_INCOME_CATEGORIES,
        SHEET_EXPENSES,
        SHEET_INCOME,
    }
    missing = [name for name in required_sheets if name not in wb.sheetnames]
    if missing:
        return [f"Missing sheets: {', '.join(missing)}"]

    accounts_rows = _sheet_as_dicts(wb[SHEET_ACCOUNTS])
    expense_cat_rows = _sheet_as_dicts(wb[SHEET_EXPENSE_CATEGORIES])
    income_cat_rows = _sheet_as_dicts(wb[SHEET_INCOME_CATEGORIES])
    expense_rows = _sheet_as_dicts(wb[SHEET_EXPENSES])
    income_rows = _sheet_as_dicts(wb[SHEET_INCOME])

    insert_expenses, insert_income, parse_errors = _collect_import_movements(
        expense_rows, income_rows
    )
    if parse_errors:
        return parse_errors

    expense_cats_from_movements = {row[2] for row in insert_expenses}
    income_cats_from_movements = {row[2] for row in insert_income}
    accounts_from_movements = {row[3] for row in insert_expenses} | {
        row[3] for row in insert_income
    }

    conn = get_connection()
    uid = int(user_id)
    try:
        conn.execute("BEGIN")

        for idx, row in enumerate(accounts_rows, start=2):
            name = row.get("name")
            if name is None or not str(name).strip():
                errors.append(f"{SHEET_ACCOUNTS} row {idx}: missing name")
                continue
            name = str(name).strip()
            opening_raw = row.get("opening_balance")
            if opening_raw is None or str(opening_raw).strip() == "":
                opening_balance = 0.0
            else:
                try:
                    opening_balance = float(opening_raw)
                except (TypeError, ValueError):
                    errors.append(f"{SHEET_ACCOUNTS} row {idx}: invalid opening_balance")
                    continue
            conn.execute(
                "INSERT OR IGNORE INTO accounts(user_id, name, opening_balance) VALUES (?, ?, ?)",
                (uid, name, opening_balance),
            )
            if sync_opening_balances:
                conn.execute(
                    "UPDATE accounts SET opening_balance = ? WHERE user_id = ? AND name = ?",
                    (opening_balance, uid, name),
                )

        if errors:
            conn.rollback()
            return errors

        for idx, row in enumerate(expense_cat_rows, start=2):
            name = row.get("name")
            if name is None or not str(name).strip():
                continue
            conn.execute(
                "INSERT OR IGNORE INTO categories(user_id, name) VALUES (?, ?)",
                (uid, str(name).strip()),
            )

        for name in sorted(expense_cats_from_movements):
            conn.execute(
                "INSERT OR IGNORE INTO categories(user_id, name) VALUES (?, ?)",
                (uid, name),
            )

        for idx, row in enumerate(income_cat_rows, start=2):
            name = row.get("name")
            if name is None or not str(name).strip():
                continue
            conn.execute(
                "INSERT OR IGNORE INTO income_categories(user_id, name) VALUES (?, ?)",
                (uid, str(name).strip()),
            )

        for name in sorted(income_cats_from_movements):
            conn.execute(
                "INSERT OR IGNORE INTO income_categories(user_id, name) VALUES (?, ?)",
                (uid, name),
            )

        for acc_name in sorted(accounts_from_movements):
            conn.execute(
                "INSERT OR IGNORE INTO accounts(user_id, name, opening_balance) VALUES (?, ?, ?)",
                (uid, acc_name, 0.0),
            )

        if replace_movements:
            conn.execute("DELETE FROM expenses WHERE user_id = ?", (uid,))
            conn.execute("DELETE FROM income_entries WHERE user_id = ?", (uid,))

        for notes, amount, cat_name, acc_name, spent_at, created_at in insert_expenses:
            category_id = _lookup_category_id(conn, cat_name, uid, expense=True)
            account_id = _lookup_account_id(conn, acc_name, uid)
            if category_id is None:
                errors.append(f"{SHEET_EXPENSES}: unknown expense category {cat_name!r}")
            if account_id is None:
                errors.append(f"{SHEET_EXPENSES}: unknown account {acc_name!r}")

        for notes, amount, cat_name, acc_name, received_at, created_at in insert_income:
            category_id = _lookup_category_id(conn, cat_name, uid, expense=False)
            account_id = _lookup_account_id(conn, acc_name, uid)
            if category_id is None:
                errors.append(f"{SHEET_INCOME}: unknown income category {cat_name!r}")
            if account_id is None:
                errors.append(f"{SHEET_INCOME}: unknown account {acc_name!r}")

        if errors:
            conn.rollback()
            return errors

        for notes, amount, cat_name, acc_name, spent_at, created_at in insert_expenses:
            category_id = _lookup_category_id(conn, cat_name, uid, expense=True)
            account_id = _lookup_account_id(conn, acc_name, uid)
            if created_at:
                conn.execute(
                    """
                    INSERT INTO expenses (user_id, notes, amount, category_id, account_id, spent_at, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (uid, notes, amount, category_id, account_id, spent_at, created_at),
                )
            else:
                conn.execute(
                    """
                    INSERT INTO expenses (user_id, notes, amount, category_id, account_id, spent_at)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (uid, notes, amount, category_id, account_id, spent_at),
                )

        for notes, amount, cat_name, acc_name, received_at, created_at in insert_income:
            category_id = _lookup_category_id(conn, cat_name, uid, expense=False)
            account_id = _lookup_account_id(conn, acc_name, uid)
            if created_at:
                conn.execute(
                    """
                    INSERT INTO income_entries (user_id, notes, amount, category_id, account_id, received_at, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (uid, notes, amount, category_id, account_id, received_at, created_at),
                )
            else:
                conn.execute(
                    """
                    INSERT INTO income_entries (user_id, notes, amount, category_id, account_id, received_at)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (uid, notes, amount, category_id, account_id, received_at),
                )

        conn.commit()
        return []
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("user_id"):
        nxt = _safe_next_url(request.args.get("next"))
        return redirect(nxt or url_for("index"))
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        conn = get_connection()
        try:
            row = conn.execute(
                "SELECT id, password_hash FROM users WHERE username = ?",
                (username,),
            ).fetchone()
        finally:
            conn.close()
        # Always run a hash comparison so response timing does not reveal
        # whether the username exists (mitigates user enumeration).
        password_hash = row["password_hash"] if row else _DUMMY_PASSWORD_HASH
        password_ok = check_password_hash(password_hash, password)
        if row and password_ok:
            session["user_id"] = row["id"]
            session["username"] = username
            nxt = _safe_next_url(request.form.get("next")) or _safe_next_url(request.args.get("next"))
            return redirect(nxt or url_for("index"))
        flash("Invalid username or password.", "error")
    return render_template("login.html", next=request.args.get("next") or "", allow_registration=ALLOW_REGISTRATION)


@app.route("/register", methods=["GET", "POST"])
def register():
    if not ALLOW_REGISTRATION:
        flash("Registration is disabled.", "error")
        return redirect(url_for("login"))
    if session.get("user_id"):
        return redirect(url_for("index"))
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        password2 = request.form.get("password2") or ""
        if not _USERNAME_RE.fullmatch(username):
            flash("Username must be 3–32 characters (letters, digits, . _ -).", "error")
            return render_template("register.html")
        if len(password) < 8:
            flash("Password must be at least 8 characters.", "error")
            return render_template("register.html")
        if password != password2:
            flash("Passwords do not match.", "error")
            return render_template("register.html")
        conn = get_connection()
        try:
            conn.execute(
                "INSERT INTO users (username, password_hash) VALUES (?, ?)",
                (username, generate_password_hash(password)),
            )
            uid = int(conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"])
            seed_user_defaults(conn, uid)
            conn.commit()
        except sqlite3.IntegrityError:
            conn.rollback()
            conn.close()
            flash("That username is already taken.", "error")
            return render_template("register.html")
        conn.close()
        session["user_id"] = uid
        session["username"] = username
        flash("Account created.", "success")
        return redirect(url_for("index"))
    return render_template("register.html")


@app.route("/logout", methods=["POST"])
def logout():
    session.pop("user_id", None)
    session.pop("username", None)
    return redirect(url_for("login"))


@app.route("/export/excel")
def export_excel():
    conn = get_connection()
    try:
        wb = _build_export_workbook(conn, g.user_id)
    finally:
        conn.close()

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    stamp = datetime.now().strftime("%Y%m%d-%H%M")
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"budget-export-{stamp}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        max_age=0,
    )


@app.route("/export/migration-template")
def export_migration_template():
    wb = _build_migration_template_workbook()
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="budget-migration-template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        max_age=0,
    )


@app.route("/import/excel", methods=["POST"])
def import_excel():
    upload = request.files.get("file")
    if not upload or upload.filename.strip() == "":
        flash("Choose an Excel file (.xlsx).", "error")
        return redirect_home(panel="settings", settings_section="migration")

    replace_movements = request.form.get("replace_movements") == "1"
    sync_opening_balances = request.form.get("sync_opening_balances") == "1"

    raw = upload.read()
    if not raw:
        flash("The uploaded file is empty.", "error")
        return redirect_home(panel="settings", settings_section="migration")

    try:
        workbook = load_workbook(BytesIO(raw), data_only=True)
    except Exception as exc:
        flash(f"Could not read the Excel file: {exc}", "error")
        return redirect_home(panel="settings", settings_section="migration")

    errors = _run_import_workbook(workbook, replace_movements, sync_opening_balances, g.user_id)
    if errors:
        preview = "; ".join(errors[:8])
        extra = f" (+{len(errors) - 8} more)" if len(errors) > 8 else ""
        flash(f"Import failed. {preview}{extra}", "error")
        return redirect_home(panel="settings", settings_section="migration")

    flash(
        "Import completed. Accounts/categories from the file were merged (new names added)."
        + (
            " Existing expense and income rows were replaced by the file."
            if replace_movements
            else " Movement rows from the file were added (existing rows kept)."
        ),
        "success",
    )
    return redirect_home(panel="settings", settings_section="migration")


@app.route("/")
def index():
    conn = get_connection()
    uid = g.user_id

    posted_recurring = apply_recurring_entries(conn, uid)
    if posted_recurring > 0:
        conn.commit()
        flash(
            f"Posted {posted_recurring} recurring entr{'ies' if posted_recurring != 1 else 'y'}.",
            "success",
        )

    raw_panel = request.args.get("panel", "").strip()
    if raw_panel == "export":
        active_panel = "settings"
        section_from_legacy = "export"
    elif raw_panel == "migration":
        active_panel = "settings"
        section_from_legacy = "migration"
    elif raw_panel in ALLOWED_PANELS:
        active_panel = raw_panel
        section_from_legacy = None
    else:
        active_panel = "home"
        section_from_legacy = None

    month_filter = resolve_month_filter_from_request()
    year_filter = normalize_year(request.args.get("year"))
    report_year = normalize_year(request.args.get("report_year"))
    year_str, month_str = month_filter.split("-", 1)
    month_start = datetime(int(year_str), int(month_str), 1)
    month_end = month_start + timedelta(days=32)
    month_end = month_end.replace(day=1)
    month_start_d = month_start.strftime("%Y-%m-%d")
    month_end_d = month_end.strftime("%Y-%m-%d")
    month_heading = f"{calendar.month_name[int(month_str)]} {year_str}"
    settings_section = normalize_settings_section(
        section_from_legacy or request.args.get("settings_section")
    )

    categories = conn.execute(
        "SELECT id, name FROM categories WHERE user_id = ? ORDER BY name",
        (uid,),
    ).fetchall()
    income_categories = conn.execute(
        "SELECT id, name FROM income_categories WHERE user_id = ? ORDER BY name",
        (uid,),
    ).fetchall()

    recurring_entries = conn.execute(
        """
        SELECT
            r.id,
            r.entry_type,
            r.amount,
            r.category_id,
            r.account_id,
            r.day_of_month,
            r.notes,
            r.enabled,
            r.created_at,
            a.name AS account_name,
            COALESCE(c.name, ic.name) AS category_name
        FROM recurring_entries r
        JOIN accounts a ON a.id = r.account_id
        LEFT JOIN categories c ON r.entry_type = 'expense' AND c.id = r.category_id
        LEFT JOIN income_categories ic ON r.entry_type = 'income' AND ic.id = r.category_id
        WHERE r.user_id = ?
        ORDER BY r.day_of_month, r.id
        """,
        (uid,),
    ).fetchall()

    expense_filter_month = resolve_list_month_filter("exp_month", "exp_cal_year", "exp_cal_month")
    expense_where = "WHERE e.user_id = ?"
    expense_where_params = [uid]
    if expense_filter_month:
        eb = month_bounds_dates(expense_filter_month)
        expense_where += " AND date(e.spent_at) >= date(?) AND date(e.spent_at) < date(?)"
        expense_where_params.extend([eb[0], eb[1]])
    expense_filter_category_id = normalize_optional_category_id(
        request.args.get("exp_category") or request.form.get("exp_category")
    )
    if expense_filter_category_id is not None:
        if not conn.execute(
            "SELECT 1 FROM categories WHERE id = ? AND user_id = ?",
            (expense_filter_category_id, uid),
        ).fetchone():
            expense_filter_category_id = None
        else:
            expense_where += " AND e.category_id = ?"
            expense_where_params.append(expense_filter_category_id)

    expense_total = conn.execute(
        f"SELECT COUNT(*) AS n FROM expenses e {expense_where}",
        expense_where_params,
    ).fetchone()["n"]
    expense_num_pages = max(1, (expense_total + LIST_PAGE_SIZE - 1) // LIST_PAGE_SIZE)
    expense_page = min(normalize_list_page(request.args.get("exp_page")), expense_num_pages)
    expense_offset = (expense_page - 1) * LIST_PAGE_SIZE
    expenses = conn.execute(
        f"""
        SELECT
            e.id,
            e.category_id,
            e.account_id,
            e.notes,
            e.amount,
            e.spent_at,
            e.created_at,
            c.name AS category_name,
            a.name AS account_name
        FROM expenses e
        JOIN categories c ON c.id = e.category_id
        JOIN accounts a ON a.id = e.account_id
        {expense_where}
        ORDER BY e.spent_at DESC
        LIMIT ? OFFSET ?
        """,
        (*expense_where_params, LIST_PAGE_SIZE, expense_offset),
    ).fetchall()

    accounts = conn.execute(
        "SELECT id, name, opening_balance FROM accounts WHERE user_id = ? ORDER BY name",
        (uid,),
    ).fetchall()

    income_filter_month = resolve_list_month_filter("inc_month", "inc_cal_year", "inc_cal_month")
    income_where = "WHERE i.user_id = ?"
    income_where_params = [uid]
    if income_filter_month:
        ib = month_bounds_dates(income_filter_month)
        income_where += " AND date(i.received_at) >= date(?) AND date(i.received_at) < date(?)"
        income_where_params.extend([ib[0], ib[1]])
    income_filter_category_id = normalize_optional_category_id(
        request.args.get("inc_category") or request.form.get("inc_category")
    )
    if income_filter_category_id is not None:
        if not conn.execute(
            "SELECT 1 FROM income_categories WHERE id = ? AND user_id = ?",
            (income_filter_category_id, uid),
        ).fetchone():
            income_filter_category_id = None
        else:
            income_where += " AND i.category_id = ?"
            income_where_params.append(income_filter_category_id)

    income_total = conn.execute(
        f"SELECT COUNT(*) AS n FROM income_entries i {income_where}",
        income_where_params,
    ).fetchone()["n"]
    income_num_pages = max(1, (income_total + LIST_PAGE_SIZE - 1) // LIST_PAGE_SIZE)
    income_page = min(normalize_list_page(request.args.get("inc_page")), income_num_pages)
    income_offset = (income_page - 1) * LIST_PAGE_SIZE
    income_entries = conn.execute(
        f"""
        SELECT
            i.id,
            i.category_id,
            i.account_id,
            i.notes,
            i.amount,
            i.received_at,
            i.created_at,
            c.name AS category_name,
            a.name AS account_name
        FROM income_entries i
        JOIN income_categories c ON c.id = i.category_id
        JOIN accounts a ON a.id = i.account_id
        {income_where}
        ORDER BY i.received_at DESC
        LIMIT ? OFFSET ?
        """,
        (*income_where_params, LIST_PAGE_SIZE, income_offset),
    ).fetchall()

    total_expenses = conn.execute(
        """
        SELECT COALESCE(-SUM(amount), 0) AS total
        FROM expenses
        WHERE user_id = ? AND date(spent_at) >= date(?) AND date(spent_at) < date(?)
        """,
        (uid, month_start_d, month_end_d),
    ).fetchone()["total"]
    total_income = conn.execute(
        """
        SELECT COALESCE(SUM(amount), 0) AS total
        FROM income_entries
        WHERE user_id = ? AND date(received_at) >= date(?) AND date(received_at) < date(?)
        """,
        (uid, month_start_d, month_end_d),
    ).fetchone()["total"]

    expense_breakdown = conn.execute(
        """
        SELECT c.name AS category_name, COALESCE(-SUM(e.amount), 0) AS total_amount
        FROM expenses e
        JOIN categories c ON c.id = e.category_id
        WHERE e.user_id = ? AND date(e.spent_at) >= date(?) AND date(e.spent_at) < date(?)
        GROUP BY c.name
        ORDER BY total_amount DESC
        """,
        (uid, month_start_d, month_end_d),
    ).fetchall()

    income_breakdown = conn.execute(
        """
        SELECT c.name AS category_name, SUM(i.amount) AS total_amount
        FROM income_entries i
        JOIN income_categories c ON c.id = i.category_id
        WHERE i.user_id = ? AND date(i.received_at) >= date(?) AND date(i.received_at) < date(?)
        GROUP BY c.name
        ORDER BY total_amount DESC
        """,
        (uid, month_start_d, month_end_d),
    ).fetchall()

    today_d = datetime.now().date()
    sel_y, sel_m = int(year_str), int(month_str)
    if (today_d.year, today_d.month) == (sel_y, sel_m):
        balance_cutoff_d = (today_d + timedelta(days=1)).isoformat()
        balance_scope_hint = (
            "Account balances use opening balance plus movements through today "
            "(only while this month is selected)."
        )
    else:
        balance_cutoff_d = month_end_d
        balance_scope_hint = (
            f"Account balances are through the end of {month_heading} "
            "(opening balance plus movements up to then)."
        )

    account_balances = fetch_account_balances_through(conn, balance_cutoff_d, uid)
    account_balance_by_id = {
        int(row["id"]): float(row["current_balance"]) for row in account_balances
    }
    account_transfers = conn.execute(
        """
        SELECT
            t.id,
            t.amount,
            t.transferred_at,
            t.notes,
            fa.name AS from_account_name,
            ta.name AS to_account_name
        FROM account_transfers t
        JOIN accounts fa ON fa.id = t.from_account_id
        JOIN accounts ta ON ta.id = t.to_account_id
        WHERE t.user_id = ?
        ORDER BY date(t.transferred_at) DESC, t.id DESC
        LIMIT ?
        """,
        (uid, TRANSFER_LOG_LIMIT),
    ).fetchall()
    accounts_total_balance = sum(float(r["current_balance"]) for r in account_balances)

    year_start_d = f"{year_filter:04d}-01-01"
    year_end_d = f"{year_filter + 1:04d}-01-01"
    expense_by_month = {
        row["m"]: float(row["total"])
        for row in conn.execute(
            """
            SELECT CAST(strftime('%m', spent_at) AS INTEGER) AS m,
                   COALESCE(-SUM(amount), 0) AS total
            FROM expenses
            WHERE user_id = ? AND date(spent_at) >= date(?) AND date(spent_at) < date(?)
            GROUP BY m
            """,
            (uid, year_start_d, year_end_d),
        )
    }
    income_by_month = {
        row["m"]: float(row["total"])
        for row in conn.execute(
            """
            SELECT CAST(strftime('%m', received_at) AS INTEGER) AS m,
                   COALESCE(SUM(amount), 0) AS total
            FROM income_entries
            WHERE user_id = ? AND date(received_at) >= date(?) AND date(received_at) < date(?)
            GROUP BY m
            """,
            (uid, year_start_d, year_end_d),
        )
    }

    month_names = calendar.month_name
    yearly_month_balances = monthly_total_balances_for_year(conn, year_filter, today_d, uid)
    yearly_rows = []
    yearly_total_income = 0.0
    yearly_total_expenses = 0.0
    for month_num in range(1, 13):
        inc = income_by_month.get(month_num, 0.0)
        exp = expense_by_month.get(month_num, 0.0)
        delta = inc - exp
        yearly_total_income += inc
        yearly_total_expenses += exp
        yearly_rows.append(
            {
                "month_label": month_names[month_num],
                "income": inc,
                "expenses": exp,
                "delta": delta,
                "total_balance": yearly_month_balances[month_num - 1],
            }
        )
    yearly_total_delta = yearly_total_income - yearly_total_expenses
    transfer_default_date = datetime.now().date().isoformat()

    investments_section = normalize_investments_section(request.args.get("investments_section"))

    crypto_txs = conn.execute(
        """
        SELECT id, coin_id, coin_symbol, coin_name, tx_type, quantity,
               price_per_unit, fee, exchange, transacted_at, notes
        FROM crypto_transactions
        WHERE user_id = ?
        ORDER BY transacted_at DESC, id DESC
        """,
        (uid,),
    ).fetchall()

    crypto_holdings_raw = compute_crypto_holdings(crypto_txs)

    force_refresh = request.args.get("refresh_prices") == "1"
    coin_ids = [h["coin_id"] for h in crypto_holdings_raw]
    crypto_prices = fetch_coingecko_prices(coin_ids, force=force_refresh) if coin_ids else {}

    crypto_total_value = 0.0
    crypto_total_invested = 0.0
    for h in crypto_holdings_raw:
        cid = h["coin_id"]
        price_info = crypto_prices.get(cid)
        if price_info:
            h["current_price"] = price_info["price"]
            h["change_24h"] = price_info.get("change_24h")
            h["current_value"] = h["quantity"] * price_info["price"]
            h["pnl"] = h["current_value"] - h["total_cost"]
            h["pnl_pct"] = (h["pnl"] / h["total_cost"] * 100) if h["total_cost"] > 0 else 0
            crypto_total_value += h["current_value"]
        else:
            h["current_price"] = None
            h["change_24h"] = None
            h["current_value"] = None
            h["pnl"] = None
            h["pnl_pct"] = 0
        crypto_total_invested += h["total_cost"]

    crypto_total_pnl = crypto_total_value - crypto_total_invested
    crypto_total_pnl_pct = (crypto_total_pnl / crypto_total_invested * 100) if crypto_total_invested > 0 else 0

    cache_age = time.time() - _price_cache["fetched_at"]
    if _price_cache["fetched_at"] > 0 and cache_age < PRICE_CACHE_TTL:
        mins = int(cache_age // 60)
        secs = int(cache_age % 60)
        crypto_prices_age = f"{mins}m {secs}s ago" if mins else f"{secs}s ago"
    else:
        crypto_prices_age = ""

    stock_txs = conn.execute(
        """
        SELECT id, symbol, ticker, instrument_name, tx_type, quantity,
               price_per_unit, fee, broker, transacted_at, notes
        FROM stock_transactions
        WHERE user_id = ?
        ORDER BY transacted_at DESC, id DESC
        """,
        (uid,),
    ).fetchall()

    stock_holdings_raw = compute_stock_holdings(stock_txs)
    stock_symbols = [h["symbol"] for h in stock_holdings_raw]
    force_stock_refresh = request.args.get("refresh_stock_prices") == "1"
    stock_prices = fetch_finnhub_quotes(stock_symbols, force=force_stock_refresh) if stock_symbols else {}

    stock_total_value = 0.0
    stock_total_invested = 0.0
    for h in stock_holdings_raw:
        sym = h["symbol"]
        price_info = stock_prices.get(sym)
        if price_info:
            h["current_price"] = price_info["price"]
            h["change_24h"] = price_info.get("change_24h")
            h["current_value"] = h["quantity"] * price_info["price"]
            h["pnl"] = h["current_value"] - h["total_cost"]
            h["pnl_pct"] = (h["pnl"] / h["total_cost"] * 100) if h["total_cost"] > 0 else 0
            stock_total_value += h["current_value"]
        else:
            h["current_price"] = None
            h["change_24h"] = None
            h["current_value"] = None
            h["pnl"] = None
            h["pnl_pct"] = 0
        stock_total_invested += h["total_cost"]

    stock_total_pnl = stock_total_value - stock_total_invested
    stock_total_pnl_pct = (stock_total_pnl / stock_total_invested * 100) if stock_total_invested > 0 else 0

    stock_cache_age = time.time() - _stock_price_cache["fetched_at"]
    if _stock_price_cache["fetched_at"] > 0 and stock_cache_age < PRICE_CACHE_TTL:
        mins = int(stock_cache_age // 60)
        secs = int(stock_cache_age % 60)
        stock_prices_age = f"{mins}m {secs}s ago" if mins else f"{secs}s ago"
    else:
        stock_prices_age = ""

    reports_section = normalize_reports_section(request.args.get("reports_section"))
    report_account_id = normalize_report_account(request.args.get("report_account"), conn, uid)
    report_account_label = "All accounts"
    if report_account_id is not None:
        for acc in accounts:
            if int(acc["id"]) == int(report_account_id):
                report_account_label = str(acc["name"])
                break

    # Reports involve month-by-month balance/portfolio math and can trigger
    # synchronous external price lookups (CoinGecko/yfinance). Only compute them
    # when the Reports panel is actually being viewed.
    report_live_balance = 0.0
    report_chart_spec = balance_line_chart_spec([])
    crypto_chart_spec = balance_line_chart_spec([])
    stock_chart_spec = balance_line_chart_spec([])
    reports_expenses_table = _empty_expense_pivot()
    if active_panel == "reports":
        jan1_this_year = f"{report_year:04d}-01-01"
        prev_balance_total = account_balance_at_cutoff(
            conn, jan1_this_year, uid, report_account_id
        )
        report_live_cutoff = (today_d + timedelta(days=1)).isoformat()
        report_live_balance = account_balance_at_cutoff(
            conn, report_live_cutoff, uid, report_account_id
        )
        report_balances_list = monthly_total_balances_for_year(
            conn, report_year, today_d, uid, report_account_id
        )
        report_balance_rows = build_monthly_chart_rows(
            report_balances_list, report_year, baseline=prev_balance_total
        )
        report_chart_spec = balance_line_chart_spec(report_balance_rows)
        reports_expenses_table = expense_pivot_for_report_year(conn, report_year, uid)

        crypto_baseline = portfolio_baseline_before_year(
            conn, crypto_txs, report_year, today_d, compute_crypto_holdings, "coin_id", "crypto"
        )
        crypto_monthly = monthly_crypto_portfolio_values_for_year(
            conn, crypto_txs, report_year, today_d
        )
        crypto_chart_rows = build_monthly_chart_rows(
            crypto_monthly, report_year, baseline=crypto_baseline
        )
        crypto_chart_spec = balance_line_chart_spec(crypto_chart_rows)

        stock_baseline = portfolio_baseline_before_year(
            conn, stock_txs, report_year, today_d, compute_stock_holdings, "symbol", "stock"
        )
        stock_monthly = monthly_stock_portfolio_values_for_year(
            conn, stock_txs, report_year, today_d
        )
        stock_chart_rows = build_monthly_chart_rows(
            stock_monthly, report_year, baseline=stock_baseline
        )
        stock_chart_spec = balance_line_chart_spec(stock_chart_rows)

    user_integrations = integrations.get_user_integrations(conn, uid)
    telegram_server = telegram_bot.server_config_for_form(conn)
    telegram_link = telegram_bot.get_telegram_link(conn, uid)
    telegram_link_code = telegram_bot.get_active_link_code(conn, uid)
    telegram_cfg = telegram_bot.get_server_config(conn)
    telegram_enabled = telegram_bot.is_configured(telegram_cfg)
    telegram_bot_username = None
    if telegram_enabled and settings_section == "integrations":
        telegram_bot_username = telegram_bot.get_bot_username(telegram_cfg)

    conn.commit()
    conn.close()

    return render_template(
        "index.html",
        session_username=g.username,
        user_integrations=user_integrations,
        telegram_server=telegram_server,
        telegram_link=telegram_link,
        telegram_link_code=telegram_link_code,
        telegram_enabled=telegram_enabled,
        telegram_bot_username=telegram_bot_username,
        categories=categories,
        income_categories=income_categories,
        expenses=expenses,
        accounts=accounts,
        income_entries=income_entries,
        active_panel=active_panel,
        month_filter=month_filter,
        year_filter=year_filter,
        settings_section=settings_section,
        total_expenses=total_expenses,
        total_income=total_income,
        net_balance=total_income - total_expenses,
        accounts_total_balance=accounts_total_balance,
        expense_breakdown=expense_breakdown,
        income_breakdown=income_breakdown,
        account_balances=account_balances,
        account_balance_by_id=account_balance_by_id,
        account_transfers=account_transfers,
        yearly_rows=yearly_rows,
        yearly_total_income=yearly_total_income,
        yearly_total_expenses=yearly_total_expenses,
        yearly_total_delta=yearly_total_delta,
        expense_page=expense_page,
        expense_total=expense_total,
        expense_num_pages=expense_num_pages,
        income_page=income_page,
        income_total=income_total,
        income_num_pages=income_num_pages,
        list_page_size=LIST_PAGE_SIZE,
        expense_filter_month=expense_filter_month,
        expense_filter_category_id=expense_filter_category_id,
        income_filter_month=income_filter_month,
        income_filter_category_id=income_filter_category_id,
        month_heading=month_heading,
        balance_scope_hint=balance_scope_hint,
        cal=calendar,
        transfer_default_date=transfer_default_date,
        transfer_log_limit=TRANSFER_LOG_LIMIT,
        report_year=report_year,
        reports_section=reports_section,
        report_account_id=report_account_id,
        report_account_label=report_account_label,
        report_live_balance=report_live_balance,
        report_today_month=calendar.month_name[today_d.month],
        report_chart_spec=report_chart_spec,
        crypto_chart_spec=crypto_chart_spec,
        stock_chart_spec=stock_chart_spec,
        reports_expenses_table=reports_expenses_table,
        recurring_entries=recurring_entries,
        investments_section=investments_section,
        crypto_holdings=crypto_holdings_raw,
        crypto_transactions=crypto_txs,
        crypto_total_value=crypto_total_value,
        crypto_total_invested=crypto_total_invested,
        crypto_total_pnl=crypto_total_pnl,
        crypto_total_pnl_pct=crypto_total_pnl_pct,
        crypto_prices_age=crypto_prices_age,
        stock_holdings=stock_holdings_raw,
        stock_transactions=stock_txs,
        stock_total_value=stock_total_value,
        stock_total_invested=stock_total_invested,
        stock_total_pnl=stock_total_pnl,
        stock_total_pnl_pct=stock_total_pnl_pct,
        stock_prices_age=stock_prices_age,
        finnhub_configured=bool(FINNHUB_API_KEY),
    )


def _parse_category_choice(raw):
    """Return ('expense'|'income', category_id) or (None, None)."""
    s = (raw or "").strip()
    if s.startswith("e-"):
        try:
            return "expense", int(s[2:])
        except ValueError:
            return None, None
    if s.startswith("i-"):
        try:
            return "income", int(s[2:])
        except ValueError:
            return None, None
    return None, None


def _recurring_category_ok(conn, entry_type, category_id, user_id):
    uid = int(user_id)
    if entry_type == "expense":
        return conn.execute(
            "SELECT 1 FROM categories WHERE id = ? AND user_id = ?",
            (category_id, uid),
        ).fetchone()
    if entry_type == "income":
        return conn.execute(
            "SELECT 1 FROM income_categories WHERE id = ? AND user_id = ?",
            (category_id, uid),
        ).fetchone()
    return None


@app.route("/recurring/add", methods=["POST"])
def add_recurring():
    uid = g.user_id
    entry_type, category_id = _parse_category_choice(request.form.get("category_choice"))
    amount_raw = request.form.get("amount", "").strip()
    account_raw = request.form.get("account_id", "").strip()
    day_raw = request.form.get("day_of_month", "1").strip()
    notes = request.form.get("notes", "").strip()

    if not entry_type or category_id is None or not account_raw:
        flash("Choose type, category, and account.", "error")
        return redirect_home(panel="recurring")

    conn = get_connection()
    if not _recurring_category_ok(conn, entry_type, category_id, uid):
        conn.close()
        flash("Invalid category for that type.", "error")
        return redirect_home(panel="recurring")

    try:
        account_id = int(account_raw)
        if not conn.execute(
            "SELECT 1 FROM accounts WHERE id = ? AND user_id = ?",
            (account_id, uid),
        ).fetchone():
            conn.close()
            flash("Invalid account.", "error")
            return redirect_home(panel="recurring")
        amt = abs(float(amount_raw))
        dom = normalize_day_of_month(day_raw)
    except (TypeError, ValueError):
        conn.close()
        flash("Invalid amount or account.", "error")
        return redirect_home(panel="recurring")

    if amt <= 0:
        conn.close()
        flash("Amount must be positive.", "error")
        return redirect_home(panel="recurring")

    conn.execute(
        """
        INSERT INTO recurring_entries (user_id, entry_type, amount, category_id, account_id, day_of_month, notes, enabled)
        VALUES (?, ?, ?, ?, ?, ?, ?, 1)
        """,
        (uid, entry_type, amt, category_id, account_id, dom, notes),
    )
    conn.commit()
    conn.close()
    flash("Recurring rule added.", "success")
    return redirect_home(panel="recurring")


@app.route("/recurring/<int:recurring_id>/edit", methods=["POST"])
def edit_recurring(recurring_id):
    uid = g.user_id
    entry_type, category_id = _parse_category_choice(request.form.get("category_choice"))
    amount_raw = request.form.get("amount", "").strip()
    account_raw = request.form.get("account_id", "").strip()
    day_raw = request.form.get("day_of_month", "1").strip()
    notes = request.form.get("notes", "").strip()
    enabled = 1 if request.form.get("enabled") == "1" else 0

    if not entry_type or category_id is None or not account_raw:
        flash("Choose type, category, and account.", "error")
        return redirect_home(panel="recurring")

    conn = get_connection()
    if not _recurring_category_ok(conn, entry_type, category_id, uid):
        conn.close()
        flash("Invalid category for that type.", "error")
        return redirect_home(panel="recurring")

    try:
        account_id = int(account_raw)
        if not conn.execute(
            "SELECT 1 FROM accounts WHERE id = ? AND user_id = ?",
            (account_id, uid),
        ).fetchone():
            conn.close()
            flash("Invalid account.", "error")
            return redirect_home(panel="recurring")
        amt = abs(float(amount_raw))
        dom = normalize_day_of_month(day_raw)
    except (TypeError, ValueError):
        conn.close()
        flash("Invalid amount or account.", "error")
        return redirect_home(panel="recurring")

    if amt <= 0:
        conn.close()
        flash("Amount must be positive.", "error")
        return redirect_home(panel="recurring")

    cur = conn.execute(
        """
        UPDATE recurring_entries
        SET entry_type = ?, amount = ?, category_id = ?, account_id = ?, day_of_month = ?, notes = ?, enabled = ?
        WHERE id = ? AND user_id = ?
        """,
        (entry_type, amt, category_id, account_id, dom, notes, enabled, recurring_id, uid),
    )
    conn.commit()
    conn.close()
    if cur.rowcount == 0:
        flash("Rule not found.", "error")
    else:
        flash("Recurring rule updated.", "success")
    return redirect_home(panel="recurring")


@app.route("/recurring/<int:recurring_id>/delete", methods=["POST"])
def delete_recurring(recurring_id):
    conn = get_connection()
    conn.execute(
        "DELETE FROM recurring_entries WHERE id = ? AND user_id = ?",
        (recurring_id, g.user_id),
    )
    conn.commit()
    conn.close()
    flash("Recurring rule removed.", "success")
    return redirect_home(panel="recurring")


def _stock_redirect():
    return redirect_home(panel="investments", settings_section="stocks")


@app.route("/stocks/search")
def search_stocks():
    query = (request.args.get("q") or "").strip()
    if len(query) < 1:
        return app.response_class(response=json.dumps([]), status=200, mimetype="application/json")
    if not FINNHUB_API_KEY:
        return app.response_class(response=json.dumps([]), status=200, mimetype="application/json")
    data = _finnhub_request(f"/search?q={quote(query)}")
    items = []
    for row in (data.get("result") if data else []) or []:
        sym = (row.get("symbol") or "").strip()
        if not sym:
            continue
        desc = (row.get("description") or "").strip()
        typ = (row.get("type") or "").strip()
        items.append(
            {
                "symbol": sym,
                "ticker": sym.split(".")[0].upper(),
                "name": desc or sym,
                "type": typ,
            }
        )
        if len(items) >= 12:
            break
    q_upper = query.upper()
    items.sort(
        key=lambda x: (
            0 if x["symbol"].upper() == q_upper else 1,
            0 if x["symbol"].upper().startswith(q_upper) else 2,
            x["symbol"],
        )
    )
    return app.response_class(response=json.dumps(items), status=200, mimetype="application/json")


@app.route("/stocks/add", methods=["POST"])
def add_stock():
    uid = g.user_id
    symbol = request.form.get("symbol", "").strip().upper()
    ticker = request.form.get("ticker", "").strip().upper()
    instrument_name = request.form.get("instrument_name", "").strip()
    tx_type = request.form.get("tx_type", "").strip().lower()
    quantity_raw = request.form.get("quantity", "").strip()
    price_raw = request.form.get("price_per_unit", "").strip()
    fee_raw = request.form.get("fee", "0").strip()
    broker = request.form.get("broker", "").strip()
    transacted_at = normalize_txn_day_from_form(request.form.get("transacted_at", "").strip())
    notes = request.form.get("notes", "").strip()

    if not symbol or not ticker or not instrument_name:
        flash("Fill in symbol, ticker, and name.", "error")
        return _stock_redirect()

    if tx_type not in ("buy", "sell"):
        flash("Invalid transaction type.", "error")
        return _stock_redirect()

    try:
        quantity = float(quantity_raw)
        price = float(price_raw)
        fee = abs(float(fee_raw)) if fee_raw else 0.0
    except (TypeError, ValueError):
        flash("Invalid quantity, price, or fee.", "error")
        return _stock_redirect()

    if quantity <= 0 or price < 0:
        flash("Quantity must be positive and price non-negative.", "error")
        return _stock_redirect()

    conn = get_connection()
    conn.execute(
        """
        INSERT INTO stock_transactions
            (user_id, symbol, ticker, instrument_name, tx_type, quantity, price_per_unit, fee, broker, transacted_at, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (uid, symbol, ticker, instrument_name, tx_type, quantity, price, fee, broker, transacted_at, notes),
    )
    conn.commit()
    conn.close()
    flash(f"Stock {tx_type} recorded.", "success")
    return _stock_redirect()


@app.route("/stocks/<int:tx_id>/edit", methods=["POST"])
def edit_stock(tx_id):
    uid = g.user_id
    symbol = request.form.get("symbol", "").strip().upper()
    ticker = request.form.get("ticker", "").strip().upper()
    instrument_name = request.form.get("instrument_name", "").strip()
    tx_type = request.form.get("tx_type", "").strip().lower()
    quantity_raw = request.form.get("quantity", "").strip()
    price_raw = request.form.get("price_per_unit", "").strip()
    fee_raw = request.form.get("fee", "0").strip()
    broker = request.form.get("broker", "").strip()
    transacted_at = normalize_txn_day_from_form(request.form.get("transacted_at", "").strip())
    notes = request.form.get("notes", "").strip()

    if not symbol or not ticker or not instrument_name:
        flash("Fill in symbol, ticker, and name.", "error")
        return _stock_redirect()

    if tx_type not in ("buy", "sell"):
        flash("Invalid transaction type.", "error")
        return _stock_redirect()

    try:
        quantity = float(quantity_raw)
        price = float(price_raw)
        fee = abs(float(fee_raw)) if fee_raw else 0.0
    except (TypeError, ValueError):
        flash("Invalid quantity, price, or fee.", "error")
        return _stock_redirect()

    if quantity <= 0 or price < 0:
        flash("Quantity must be positive and price non-negative.", "error")
        return _stock_redirect()

    conn = get_connection()
    cur = conn.execute(
        """
        UPDATE stock_transactions
        SET symbol = ?, ticker = ?, instrument_name = ?, tx_type = ?,
            quantity = ?, price_per_unit = ?, fee = ?, broker = ?,
            transacted_at = ?, notes = ?
        WHERE id = ? AND user_id = ?
        """,
        (symbol, ticker, instrument_name, tx_type, quantity, price, fee, broker, transacted_at, notes, tx_id, uid),
    )
    conn.commit()
    conn.close()
    if cur.rowcount == 0:
        flash("Transaction not found.", "error")
    else:
        flash("Transaction updated.", "success")
    return _stock_redirect()


@app.route("/stocks/<int:tx_id>/delete", methods=["POST"])
def delete_stock(tx_id):
    conn = get_connection()
    conn.execute(
        "DELETE FROM stock_transactions WHERE id = ? AND user_id = ?",
        (tx_id, g.user_id),
    )
    conn.commit()
    conn.close()
    flash("Transaction removed.", "success")
    return _stock_redirect()


@app.route("/crypto/search")
def search_crypto():
    query = (request.args.get("q") or "").strip()
    if len(query) < 2:
        return app.response_class(
            response=json.dumps([]),
            status=200,
            mimetype="application/json",
        )
    url = f"https://api.coingecko.com/api/v3/search?query={query}"
    try:
        req = Request(url, headers={"Accept": "application/json", "User-Agent": "VibeBudgeting/1.0"})
        with urlopen(req, timeout=8) as resp:
            data = json.loads(resp.read().decode())
        coins = [
            {"id": c["id"], "symbol": c["symbol"], "name": c["name"], "thumb": c.get("thumb", "")}
            for c in (data.get("coins") or [])[:10]
        ]
    except (URLError, HTTPError, json.JSONDecodeError, OSError):
        coins = []
    return app.response_class(
        response=json.dumps(coins),
        status=200,
        mimetype="application/json",
    )


@app.route("/crypto/add", methods=["POST"])
def add_crypto():
    uid = g.user_id
    coin_id = request.form.get("coin_id", "").strip().lower()
    coin_symbol = request.form.get("coin_symbol", "").strip().upper()
    coin_name = request.form.get("coin_name", "").strip()
    tx_type = request.form.get("tx_type", "").strip().lower()
    quantity_raw = request.form.get("quantity", "").strip()
    price_raw = request.form.get("price_per_unit", "").strip()
    fee_raw = request.form.get("fee", "0").strip()
    exchange = request.form.get("exchange", "").strip()
    transacted_at = normalize_txn_day_from_form(request.form.get("transacted_at", "").strip())
    notes = request.form.get("notes", "").strip()

    if not coin_id or not coin_symbol or not coin_name:
        flash("Fill in Coin ID, Symbol, and Name.", "error")
        return redirect_home(panel="investments", settings_section="crypto")

    if tx_type not in ("buy", "sell"):
        flash("Invalid transaction type.", "error")
        return redirect_home(panel="investments", settings_section="crypto")

    try:
        quantity = float(quantity_raw)
        price = float(price_raw)
        fee = abs(float(fee_raw)) if fee_raw else 0.0
    except (TypeError, ValueError):
        flash("Invalid quantity, price, or fee.", "error")
        return redirect_home(panel="investments", settings_section="crypto")

    if quantity <= 0 or price < 0:
        flash("Quantity must be positive and price non-negative.", "error")
        return redirect_home(panel="investments", settings_section="crypto")

    conn = get_connection()
    conn.execute(
        """
        INSERT INTO crypto_transactions
            (user_id, coin_id, coin_symbol, coin_name, tx_type, quantity, price_per_unit, fee, exchange, transacted_at, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (uid, coin_id, coin_symbol, coin_name, tx_type, quantity, price, fee, exchange, transacted_at, notes),
    )
    conn.commit()
    conn.close()
    flash(f"Crypto {tx_type} recorded.", "success")
    return redirect_home(panel="investments", settings_section="crypto")


@app.route("/crypto/<int:tx_id>/edit", methods=["POST"])
def edit_crypto(tx_id):
    uid = g.user_id
    coin_id = request.form.get("coin_id", "").strip().lower()
    coin_symbol = request.form.get("coin_symbol", "").strip().upper()
    coin_name = request.form.get("coin_name", "").strip()
    tx_type = request.form.get("tx_type", "").strip().lower()
    quantity_raw = request.form.get("quantity", "").strip()
    price_raw = request.form.get("price_per_unit", "").strip()
    fee_raw = request.form.get("fee", "0").strip()
    exchange = request.form.get("exchange", "").strip()
    transacted_at = normalize_txn_day_from_form(request.form.get("transacted_at", "").strip())
    notes = request.form.get("notes", "").strip()

    if not coin_id or not coin_symbol or not coin_name:
        flash("Fill in Coin ID, Symbol, and Name.", "error")
        return redirect_home(panel="investments", settings_section="crypto")

    if tx_type not in ("buy", "sell"):
        flash("Invalid transaction type.", "error")
        return redirect_home(panel="investments", settings_section="crypto")

    try:
        quantity = float(quantity_raw)
        price = float(price_raw)
        fee = abs(float(fee_raw)) if fee_raw else 0.0
    except (TypeError, ValueError):
        flash("Invalid quantity, price, or fee.", "error")
        return redirect_home(panel="investments", settings_section="crypto")

    if quantity <= 0 or price < 0:
        flash("Quantity must be positive and price non-negative.", "error")
        return redirect_home(panel="investments", settings_section="crypto")

    conn = get_connection()
    cur = conn.execute(
        """
        UPDATE crypto_transactions
        SET coin_id = ?, coin_symbol = ?, coin_name = ?, tx_type = ?,
            quantity = ?, price_per_unit = ?, fee = ?, exchange = ?,
            transacted_at = ?, notes = ?
        WHERE id = ? AND user_id = ?
        """,
        (coin_id, coin_symbol, coin_name, tx_type, quantity, price, fee, exchange, transacted_at, notes, tx_id, uid),
    )
    conn.commit()
    conn.close()
    if cur.rowcount == 0:
        flash("Transaction not found.", "error")
    else:
        flash("Transaction updated.", "success")
    return redirect_home(panel="investments", settings_section="crypto")


@app.route("/crypto/<int:tx_id>/delete", methods=["POST"])
def delete_crypto(tx_id):
    conn = get_connection()
    conn.execute(
        "DELETE FROM crypto_transactions WHERE id = ? AND user_id = ?",
        (tx_id, g.user_id),
    )
    conn.commit()
    conn.close()
    flash("Transaction removed.", "success")
    return redirect_home(panel="investments", settings_section="crypto")


@app.route("/categories/add", methods=["POST"])
def add_category():
    name = request.form.get("name", "").strip()
    if name:
        conn = get_connection()
        conn.execute(
            "INSERT OR IGNORE INTO categories(user_id, name) VALUES (?, ?)",
            (g.user_id, name),
        )
        conn.commit()
        conn.close()
    return redirect_home()


@app.route("/categories/<int:category_id>/edit", methods=["POST"])
def edit_category(category_id):
    name = request.form.get("name", "").strip()
    if name:
        conn = get_connection()
        conn.execute(
            "UPDATE categories SET name = ? WHERE id = ? AND user_id = ?",
            (name, category_id, g.user_id),
        )
        conn.commit()
        conn.close()
    return redirect_home()


@app.route("/categories/<int:category_id>/delete", methods=["POST"])
def delete_category(category_id):
    conn = get_connection()
    cursor = conn.execute(
        "DELETE FROM categories WHERE id = ? AND user_id = ?",
        (category_id, g.user_id),
    )
    conn.commit()
    conn.close()
    if cursor.rowcount == 0:
        print(f"[vibe-budgeting] refused category delete {category_id} (still referenced or missing)")
    return redirect_home()


@app.route("/income-categories/add", methods=["POST"])
def add_income_category():
    name = request.form.get("name", "").strip()
    if name:
        conn = get_connection()
        conn.execute(
            "INSERT OR IGNORE INTO income_categories(user_id, name) VALUES (?, ?)",
            (g.user_id, name),
        )
        conn.commit()
        conn.close()
    return redirect_home()


@app.route("/income-categories/<int:category_id>/edit", methods=["POST"])
def edit_income_category(category_id):
    name = request.form.get("name", "").strip()
    if name:
        conn = get_connection()
        conn.execute(
            "UPDATE income_categories SET name = ? WHERE id = ? AND user_id = ?",
            (name, category_id, g.user_id),
        )
        conn.commit()
        conn.close()
    return redirect_home()


@app.route("/income-categories/<int:category_id>/delete", methods=["POST"])
def delete_income_category(category_id):
    conn = get_connection()
    cursor = conn.execute(
        "DELETE FROM income_categories WHERE id = ? AND user_id = ?",
        (category_id, g.user_id),
    )
    conn.commit()
    conn.close()
    if cursor.rowcount == 0:
        print(f"[vibe-budgeting] refused income category delete {category_id} (still referenced or missing)")
    return redirect_home()


@app.route("/expenses/add", methods=["POST"])
def add_expense():
    notes = request.form.get("notes", "").strip()
    amount = request.form.get("amount", "0").strip()
    category_raw = request.form.get("category_id", "").strip()
    account_raw = request.form.get("account_id", "").strip()

    if category_raw and account_raw:
        try:
            category_id = int(category_raw)
            account_id = int(account_raw)
        except (TypeError, ValueError):
            flash("Invalid category or account.", "error")
            return redirect_home()
        spent_at = datetime.now().date().isoformat()
        conn = get_connection()
        if not _user_owns_category(conn, category_id, g.user_id, expense=True) or not _user_owns_account(
            conn, account_id, g.user_id
        ):
            conn.close()
            flash("Invalid category or account.", "error")
            return redirect_home()
        conn.execute(
            """
            INSERT INTO expenses (user_id, notes, amount, category_id, account_id, spent_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                g.user_id,
                notes,
                normalize_expense_amount(amount),
                category_id,
                account_id,
                spent_at,
            ),
        )
        conn.commit()
        conn.close()

    return redirect_home()


@app.route("/expenses/<int:expense_id>/edit", methods=["POST"])
def edit_expense(expense_id):
    notes = request.form.get("notes", "").strip()
    amount = request.form.get("amount", "0").strip()
    category_raw = request.form.get("category_id", "").strip()
    account_raw = request.form.get("account_id", "").strip()
    spent_at_raw = request.form.get("spent_at", "").strip()

    if category_raw and account_raw:
        try:
            category_id = int(category_raw)
            account_id = int(account_raw)
        except (TypeError, ValueError):
            flash("Invalid category or account.", "error")
            return redirect_home()
        spent_at = normalize_txn_day_from_form(spent_at_raw)
        conn = get_connection()
        if not _user_owns_category(conn, category_id, g.user_id, expense=True) or not _user_owns_account(
            conn, account_id, g.user_id
        ):
            conn.close()
            flash("Invalid category or account.", "error")
            return redirect_home()
        conn.execute(
            """
            UPDATE expenses
            SET notes = ?, amount = ?, category_id = ?, account_id = ?, spent_at = ?
            WHERE id = ? AND user_id = ?
            """,
            (
                notes,
                normalize_expense_amount(amount),
                category_id,
                account_id,
                spent_at,
                expense_id,
                g.user_id,
            ),
        )
        conn.commit()
        conn.close()

    return redirect_home()


@app.route("/expenses/<int:expense_id>/delete", methods=["POST"])
def delete_expense(expense_id):
    conn = get_connection()
    conn.execute(
        "DELETE FROM expenses WHERE id = ? AND user_id = ?",
        (expense_id, g.user_id),
    )
    conn.commit()
    conn.close()
    return redirect_home()


def _parse_opening_balance(raw):
    text = (raw or "0").strip() or "0"
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


@app.route("/accounts/add", methods=["POST"])
def add_account():
    name = request.form.get("name", "").strip()
    opening_balance = _parse_opening_balance(request.form.get("opening_balance"))
    if opening_balance is None:
        flash("Invalid opening balance.", "error")
        return redirect_home()
    if name:
        conn = get_connection()
        conn.execute(
            "INSERT OR IGNORE INTO accounts(user_id, name, opening_balance) VALUES (?, ?, ?)",
            (g.user_id, name, opening_balance),
        )
        conn.commit()
        conn.close()
    return redirect_home()


@app.route("/accounts/<int:account_id>/delete", methods=["POST"])
def delete_account(account_id):
    conn = get_connection()
    cursor = conn.execute(
        "DELETE FROM accounts WHERE id = ? AND user_id = ?",
        (account_id, g.user_id),
    )
    conn.commit()
    conn.close()
    if cursor.rowcount == 0:
        print(f"[vibe-budgeting] refused account delete {account_id} (still referenced or missing)")
    return redirect_home()


@app.route("/accounts/<int:account_id>/edit", methods=["POST"])
def edit_account(account_id):
    name = request.form.get("name", "").strip()
    opening_balance = _parse_opening_balance(request.form.get("opening_balance"))
    if opening_balance is None:
        flash("Invalid opening balance.", "error")
        return redirect_home()
    if name:
        conn = get_connection()
        conn.execute(
            "UPDATE accounts SET name = ?, opening_balance = ? WHERE id = ? AND user_id = ?",
            (name, opening_balance, account_id, g.user_id),
        )
        conn.commit()
        conn.close()
    return redirect_home()


@app.route("/income/<int:income_id>/delete", methods=["POST"])
def delete_income(income_id):
    conn = get_connection()
    conn.execute(
        "DELETE FROM income_entries WHERE id = ? AND user_id = ?",
        (income_id, g.user_id),
    )
    conn.commit()
    conn.close()
    return redirect_home()


@app.route("/income/add", methods=["POST"])
def add_income():
    notes = request.form.get("notes", "").strip()
    amount = request.form.get("amount", "0").strip()
    account_raw = request.form.get("account_id", "").strip()
    category_raw = request.form.get("category_id", "").strip()

    if category_raw and account_raw:
        try:
            category_id = int(category_raw)
            account_id = int(account_raw)
        except (TypeError, ValueError):
            flash("Invalid category or account.", "error")
            return redirect_home()
        received_at = datetime.now().date().isoformat()
        conn = get_connection()
        if not _user_owns_category(conn, category_id, g.user_id, expense=False) or not _user_owns_account(
            conn, account_id, g.user_id
        ):
            conn.close()
            flash("Invalid category or account.", "error")
            return redirect_home()
        conn.execute(
            """
            INSERT INTO income_entries (user_id, notes, amount, category_id, account_id, received_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                g.user_id,
                notes,
                normalize_income_amount(amount),
                category_id,
                account_id,
                received_at,
            ),
        )
        conn.commit()
        conn.close()
    return redirect_home()


@app.route("/income/<int:income_id>/edit", methods=["POST"])
def edit_income(income_id):
    notes = request.form.get("notes", "").strip()
    amount = request.form.get("amount", "0").strip()
    account_raw = request.form.get("account_id", "").strip()
    category_raw = request.form.get("category_id", "").strip()
    received_at_raw = request.form.get("received_at", "").strip()

    if category_raw and account_raw:
        try:
            category_id = int(category_raw)
            account_id = int(account_raw)
        except (TypeError, ValueError):
            flash("Invalid category or account.", "error")
            return redirect_home()
        received_at = normalize_txn_day_from_form(received_at_raw)
        conn = get_connection()
        if not _user_owns_category(conn, category_id, g.user_id, expense=False) or not _user_owns_account(
            conn, account_id, g.user_id
        ):
            conn.close()
            flash("Invalid category or account.", "error")
            return redirect_home()
        conn.execute(
            """
            UPDATE income_entries
            SET notes = ?, amount = ?, category_id = ?, account_id = ?, received_at = ?
            WHERE id = ? AND user_id = ?
            """,
            (
                notes,
                normalize_income_amount(amount),
                category_id,
                account_id,
                received_at,
                income_id,
                g.user_id,
            ),
        )
        conn.commit()
        conn.close()

    return redirect_home()


@app.route("/transfers/add", methods=["POST"])
def add_transfer():
    from_raw = request.form.get("from_account_id", "").strip()
    to_raw = request.form.get("to_account_id", "").strip()
    amount_raw = request.form.get("amount", "").strip()
    notes = request.form.get("notes", "").strip()
    transferred_at = normalize_txn_day_from_form(request.form.get("transferred_at", "").strip())

    if not from_raw or not to_raw:
        return redirect_home()

    try:
        from_id = int(from_raw)
        to_id = int(to_raw)
    except ValueError:
        flash("Invalid accounts.", "error")
        return redirect_home()

    if from_id == to_id:
        flash("Choose two different accounts.", "error")
        return redirect_home()

    try:
        amt = abs(float(amount_raw))
    except (TypeError, ValueError):
        amt = 0.0

    if amt <= 0:
        flash("Enter a positive amount.", "error")
        return redirect_home()

    conn = get_connection()
    uid = g.user_id
    if not conn.execute(
        "SELECT 1 FROM accounts WHERE id = ? AND user_id = ?",
        (from_id, uid),
    ).fetchone() or not conn.execute(
        "SELECT 1 FROM accounts WHERE id = ? AND user_id = ?",
        (to_id, uid),
    ).fetchone():
        conn.close()
        flash("Invalid accounts.", "error")
        return redirect_home()

    conn.execute(
        """
        INSERT INTO account_transfers (user_id, from_account_id, to_account_id, amount, transferred_at, notes)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (uid, from_id, to_id, amt, transferred_at, notes),
    )
    conn.commit()
    conn.close()
    flash("Transfer recorded.", "success")
    return redirect_home()


@app.route("/transfers/<int:transfer_id>/delete", methods=["POST"])
def delete_transfer(transfer_id):
    conn = get_connection()
    cursor = conn.execute(
        "DELETE FROM account_transfers WHERE id = ? AND user_id = ?",
        (transfer_id, g.user_id),
    )
    conn.commit()
    conn.close()
    if cursor.rowcount == 0:
        flash("Transfer not found.", "error")
    return redirect_home()


def _integrations_from_request():
    try:
        return integrations.parse_integration_form(request.form)
    except ValueError as exc:
        flash(str(exc), "error")
        return None


@app.route("/settings/integrations/save", methods=["POST"])
def save_integrations():
    settings = _integrations_from_request()
    if settings is None:
        return redirect_home(panel="settings", settings_section="integrations")
    conn = get_connection()
    integrations.save_user_integrations(conn, g.user_id, settings)
    conn.close()
    flash("Integration settings saved.", "success")
    return redirect_home(panel="settings", settings_section="integrations")


@app.route("/settings/integrations/test", methods=["POST"])
def test_integrations():
    settings = _integrations_from_request()
    if settings is None:
        return redirect_home(panel="settings", settings_section="integrations")
    ok, message = integrations.test_ai_connection(settings)
    flash(message, "success" if ok else "error")
    return redirect_home(panel="settings", settings_section="integrations")


@app.route("/settings/integrations/models")
def list_integration_models():
    base_url = (request.args.get("base_url") or "").strip().rstrip("/")
    if not base_url:
        return app.response_class(
            response=json.dumps({"error": "Base URL is required."}),
            status=400,
            mimetype="application/json",
        )
    try:
        integrations._normalize_base_url(base_url)
    except ValueError as exc:
        return app.response_class(
            response=json.dumps({"error": str(exc)}),
            status=400,
            mimetype="application/json",
        )
    timeout = integrations.DEFAULT_AI_TIMEOUT
    try:
        models = integrations.fetch_ollama_models(base_url, timeout=timeout)
    except HTTPError as exc:
        return app.response_class(
            response=json.dumps({"error": f"Ollama responded with HTTP {exc.code}."}),
            status=502,
            mimetype="application/json",
        )
    except URLError as exc:
        return app.response_class(
            response=json.dumps({"error": f"Could not reach Ollama: {exc.reason}"}),
            status=502,
            mimetype="application/json",
        )
    except (json.JSONDecodeError, OSError, TimeoutError) as exc:
        return app.response_class(
            response=json.dumps({"error": f"Connection failed: {exc}"}),
            status=502,
            mimetype="application/json",
        )
    return app.response_class(
        response=json.dumps({"models": models}),
        status=200,
        mimetype="application/json",
    )


@app.route("/settings/telegram/server", methods=["POST"])
def save_telegram_server():
    conn = get_connection()
    existing = telegram_bot.get_server_config(conn)
    try:
        settings = telegram_bot.parse_server_config_form(request.form, existing)
    except ValueError as exc:
        conn.close()
        flash(str(exc), "error")
        return redirect_home(panel="settings", settings_section="integrations")
    telegram_bot.save_server_config(conn, settings)
    ok, message = telegram_bot.clear_webhook(settings)
    conn.close()
    flash("Telegram settings saved. Polling mode — no public URL needed.", "success")
    flash(message, "success" if ok else "error")
    return redirect_home(panel="settings", settings_section="integrations")


@app.route("/settings/telegram/test", methods=["POST"])
def test_telegram_server():
    conn = get_connection()
    existing = telegram_bot.get_server_config(conn)
    try:
        settings = telegram_bot.parse_server_config_form(request.form, existing)
    except ValueError as exc:
        conn.close()
        flash(str(exc), "error")
        return redirect_home(panel="settings", settings_section="integrations")
    ok, message = telegram_bot.test_telegram_connection(settings)
    conn.close()
    flash(message, "success" if ok else "error")
    return redirect_home(panel="settings", settings_section="integrations")


@app.route("/settings/telegram/generate-code", methods=["POST"])
def telegram_generate_code():
    conn = get_connection()
    config = telegram_bot.get_server_config(conn)
    if not telegram_bot.is_configured(config):
        conn.close()
        flash("Configure the Telegram bot below first.", "error")
        return redirect_home(panel="settings", settings_section="integrations")
    code = telegram_bot.create_link_code(conn, g.user_id)
    conn.close()
    flash(f"Link code: {code} (15 min). In Telegram send: /link {code}", "success")
    return redirect_home(panel="settings", settings_section="integrations")


@app.route("/settings/telegram/unlink", methods=["POST"])
def telegram_unlink():
    conn = get_connection()
    telegram_bot.unlink_telegram(conn, g.user_id)
    conn.close()
    flash("Telegram unlinked.", "success")
    return redirect_home(panel="settings", settings_section="integrations")


@app.route("/settings/telegram/default-account", methods=["POST"])
def telegram_default_account():
    raw = request.form.get("default_account_id", "").strip()
    if raw:
        try:
            account_id = int(raw)
        except (TypeError, ValueError):
            flash("Could not update default account.", "error")
            return redirect_home(panel="settings", settings_section="integrations")
    else:
        account_id = None
    conn = get_connection()
    if not telegram_bot.set_default_account(conn, g.user_id, account_id):
        conn.close()
        flash("Could not update default account.", "error")
        return redirect_home(panel="settings", settings_section="integrations")
    conn.close()
    flash("Telegram default account updated.", "success")
    return redirect_home(panel="settings", settings_section="integrations")


def arm_telegram_poller() -> None:
    global _telegram_poller_armed
    if _telegram_poller_armed:
        return
    with _telegram_poller_guard:
        if _telegram_poller_armed:
            return
        telegram_bot.start_poller(get_connection, fetch_account_balances_through, _poll_lock_path)
        _telegram_poller_armed = True


_prepare_sqlite_storage()
init_db()

_poll_lock_path = os.path.join(os.path.dirname(DB_PATH) or ".", "telegram_poll.lock")
_telegram_poller_armed = False
_telegram_poller_guard = threading.Lock()

_conn_boot = get_connection()
_boot_cfg = telegram_bot.get_server_config(_conn_boot)
if telegram_bot.is_configured(_boot_cfg):
    ok, msg = telegram_bot.clear_webhook(_boot_cfg)
    print(f"Telegram: {msg}", file=sys.stderr)
_conn_boot.close()

if __name__ == "__main__":
    arm_telegram_poller()
    app.run(host="0.0.0.0", port=5000)