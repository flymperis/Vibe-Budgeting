import re
from datetime import datetime, timedelta, timezone
from openpyxl import Workbook, load_workbook

from config import EXPORT_FORMAT_VERSION, SHEET_ACCOUNTS, SHEET_CRYPTO, SHEET_EXPENSES, SHEET_EXPENSE_CATEGORIES, SHEET_INCOME, SHEET_INCOME_CATEGORIES, SHEET_META, SHEET_RECURRING, SHEET_STOCKS, SHEET_TRANSFERS
from db import get_connection
from helpers import _lookup_account_id, _lookup_category_id

# Column order for the sheets added in format version 2. Used for both the
# export and the header-only migration template so the two cannot drift.
TRANSFER_COLUMNS = [
    "from_account_name",
    "to_account_name",
    "amount",
    "transferred_at",
    "notes",
    "created_at",
]
STOCK_COLUMNS = [
    "symbol",
    "ticker",
    "instrument_name",
    "tx_type",
    "quantity",
    "price_per_unit",
    "fee",
    "broker",
    "transacted_at",
    "notes",
    "created_at",
]
CRYPTO_COLUMNS = [
    "coin_id",
    "coin_symbol",
    "coin_name",
    "tx_type",
    "quantity",
    "price_per_unit",
    "fee",
    "exchange",
    "transacted_at",
    "notes",
    "created_at",
]
RECURRING_COLUMNS = [
    "entry_type",
    "amount",
    "category_name",
    "account_name",
    "day_of_month",
    "months_to_run",
    "notes",
    "enabled",
    "created_at",
    "applied_months",
]


def _normalize_header_key(value):
    if value is None:
        return ""
    return str(value).strip().lower()

def _sheet_as_dicts(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers_raw = [_normalize_header_key(cell) for cell in rows[0]]
    headers = []
    for raw in headers_raw:
        headers.append(raw if raw else "")
    out = []
    for row in rows[1:]:
        if row is None:
            continue
        cells = list(row)
        if not cells:
            continue
        if all(cell is None or str(cell).strip() == "" for cell in cells):
            continue
        row_dict = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row_dict[header] = cells[idx] if idx < len(cells) else None
        out.append(row_dict)
    return out

def _parse_excel_expense_amount(value, sheet, row_num):
    """Excel import: sign is kept. Negative = spending; positive = refund/credit (adds back to balance)."""
    if value is None or (isinstance(value, str) and not value.strip()):
        raise ValueError(f"{sheet} row {row_num}: missing amount")
    try:
        amount = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{sheet} row {row_num}: invalid amount") from exc
    return amount

def _parse_excel_income_amount(value, sheet, row_num):
    if value is None or (isinstance(value, str) and not value.strip()):
        raise ValueError(f"{sheet} row {row_num}: missing amount")
    try:
        amount = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{sheet} row {row_num}: invalid amount") from exc
    return abs(amount)

def _parse_excel_datetime(value, sheet, row_num, column_label):
    if value is None or (isinstance(value, str) and not value.strip()):
        raise ValueError(f"{sheet} row {row_num}: missing {column_label}")
    if isinstance(value, datetime):
        return value.replace(microsecond=0)
    text = str(value).strip()
    try:
        normalized = text.replace(" ", "T", 1)
        if len(normalized) == 10:
            normalized = f"{normalized}T00:00:00"
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        parsed = None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                parsed = datetime.strptime(text, fmt)
                break
            except ValueError:
                continue
        if parsed is None:
            raise ValueError(f"{sheet} row {row_num}: invalid {column_label}")
    return parsed.replace(microsecond=0)

def _parse_excel_timestamp(value, sheet, row_num, column_label):
    return _parse_excel_datetime(value, sheet, row_num, column_label).isoformat(timespec="seconds")

def _parse_excel_movement_date(value, sheet, row_num, column_label):
    return _parse_excel_datetime(value, sheet, row_num, column_label).date().isoformat()

def _optional_created_at(value, sheet, row_num):
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    return _parse_excel_timestamp(value, sheet, row_num, "created_at")

def _build_export_workbook(conn, user_id):
    uid = int(user_id)
    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = SHEET_META
    ws_meta.append(["key", "value"])
    ws_meta.append(["format_version", EXPORT_FORMAT_VERSION])
    ws_meta.append(["exported_at", datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")])

    ws_accounts = wb.create_sheet(SHEET_ACCOUNTS)
    ws_accounts.append(["name", "opening_balance"])
    for row in conn.execute(
        "SELECT name, opening_balance FROM accounts WHERE user_id = ? ORDER BY name",
        (uid,),
    ):
        ws_accounts.append([row["name"], row["opening_balance"]])

    ws_ec = wb.create_sheet(SHEET_EXPENSE_CATEGORIES)
    ws_ec.append(["name"])
    for row in conn.execute(
        "SELECT name FROM categories WHERE user_id = ? ORDER BY name",
        (uid,),
    ):
        ws_ec.append([row["name"]])

    ws_ic = wb.create_sheet(SHEET_INCOME_CATEGORIES)
    ws_ic.append(["name"])
    for row in conn.execute(
        "SELECT name FROM income_categories WHERE user_id = ? ORDER BY name",
        (uid,),
    ):
        ws_ic.append([row["name"]])

    ws_exp = wb.create_sheet(SHEET_EXPENSES)
    ws_exp.append(["notes", "amount", "category_name", "account_name", "spent_at", "created_at"])
    for row in conn.execute(
        """
        SELECT e.notes, e.amount, c.name AS category_name, a.name AS account_name, e.spent_at, e.created_at
        FROM expenses e
        JOIN categories c ON c.id = e.category_id
        JOIN accounts a ON a.id = e.account_id
        WHERE e.user_id = ?
        ORDER BY e.spent_at ASC, e.id ASC
        """,
        (uid,),
    ):
        ws_exp.append(
            [
                row["notes"],
                row["amount"],
                row["category_name"],
                row["account_name"],
                row["spent_at"],
                row["created_at"],
            ]
        )

    ws_inc = wb.create_sheet(SHEET_INCOME)
    ws_inc.append(["notes", "amount", "category_name", "account_name", "received_at", "created_at"])
    for row in conn.execute(
        """
        SELECT i.notes, i.amount, c.name AS category_name, a.name AS account_name, i.received_at, i.created_at
        FROM income_entries i
        JOIN income_categories c ON c.id = i.category_id
        JOIN accounts a ON a.id = i.account_id
        WHERE i.user_id = ?
        ORDER BY i.received_at ASC, i.id ASC
        """,
        (uid,),
    ):
        ws_inc.append(
            [
                row["notes"],
                row["amount"],
                row["category_name"],
                row["account_name"],
                row["received_at"],
                row["created_at"],
            ]
        )

    ws_tr = wb.create_sheet(SHEET_TRANSFERS)
    ws_tr.append(TRANSFER_COLUMNS)
    for row in conn.execute(
        """
        SELECT fa.name AS from_account_name, ta.name AS to_account_name,
               t.amount, t.transferred_at, t.notes, t.created_at
        FROM account_transfers t
        JOIN accounts fa ON fa.id = t.from_account_id
        JOIN accounts ta ON ta.id = t.to_account_id
        WHERE t.user_id = ?
        ORDER BY t.transferred_at ASC, t.id ASC
        """,
        (uid,),
    ):
        ws_tr.append([row[c] for c in TRANSFER_COLUMNS])

    ws_st = wb.create_sheet(SHEET_STOCKS)
    ws_st.append(STOCK_COLUMNS)
    for row in conn.execute(
        """
        SELECT symbol, ticker, instrument_name, tx_type, quantity, price_per_unit,
               fee, broker, transacted_at, notes, created_at
        FROM stock_transactions
        WHERE user_id = ?
        ORDER BY transacted_at ASC, id ASC
        """,
        (uid,),
    ):
        ws_st.append([row[c] for c in STOCK_COLUMNS])

    ws_cr = wb.create_sheet(SHEET_CRYPTO)
    ws_cr.append(CRYPTO_COLUMNS)
    for row in conn.execute(
        """
        SELECT coin_id, coin_symbol, coin_name, tx_type, quantity, price_per_unit,
               fee, exchange, transacted_at, notes, created_at
        FROM crypto_transactions
        WHERE user_id = ?
        ORDER BY transacted_at ASC, id ASC
        """,
        (uid,),
    ):
        ws_cr.append([row[c] for c in CRYPTO_COLUMNS])

    ws_rec = wb.create_sheet(SHEET_RECURRING)
    ws_rec.append(RECURRING_COLUMNS)
    for row in conn.execute(
        """
        SELECT r.id, r.entry_type, r.amount,
               COALESCE(c.name, ic.name) AS category_name,
               a.name AS account_name, r.day_of_month, r.months_to_run,
               r.notes, r.enabled, r.created_at
        FROM recurring_entries r
        JOIN accounts a ON a.id = r.account_id
        LEFT JOIN categories c ON r.entry_type = 'expense' AND c.id = r.category_id
        LEFT JOIN income_categories ic ON r.entry_type = 'income' AND ic.id = r.category_id
        WHERE r.user_id = ?
        ORDER BY r.id ASC
        """,
        (uid,),
    ):
        # The months a rule has already posted travel with it. Without them,
        # apply_recurring_entries would re-post every month back to created_at
        # on the first page load after an import, duplicating the rows the
        # Expenses/Income sheets just restored.
        applied = [
            str(m["ym"])
            for m in conn.execute(
                "SELECT ym FROM recurring_applied WHERE recurring_id = ? ORDER BY ym",
                (row["id"],),
            )
        ]
        values = [row[c] for c in RECURRING_COLUMNS if c != "applied_months"]
        ws_rec.append(values + [",".join(applied)])

    return wb

def _build_migration_template_workbook():
    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = SHEET_META
    ws_meta.append(["key", "value"])
    ws_meta.append(["format_version", EXPORT_FORMAT_VERSION])
    ws_meta.append(["kind", "migration_template"])

    ws_accounts = wb.create_sheet(SHEET_ACCOUNTS)
    ws_accounts.append(["name", "opening_balance"])
    ws_accounts.append(["Main", 0])

    ws_ec = wb.create_sheet(SHEET_EXPENSE_CATEGORIES)
    ws_ec.append(["name"])
    ws_ec.append(["General"])

    ws_ic = wb.create_sheet(SHEET_INCOME_CATEGORIES)
    ws_ic.append(["name"])
    ws_ic.append(["General"])

    ws_exp = wb.create_sheet(SHEET_EXPENSES)
    ws_exp.append(["notes", "amount", "category_name", "account_name", "spent_at", "created_at"])

    ws_inc = wb.create_sheet(SHEET_INCOME)
    ws_inc.append(["notes", "amount", "category_name", "account_name", "received_at", "created_at"])

    wb.create_sheet(SHEET_TRANSFERS).append(TRANSFER_COLUMNS)
    wb.create_sheet(SHEET_STOCKS).append(STOCK_COLUMNS)
    wb.create_sheet(SHEET_CRYPTO).append(CRYPTO_COLUMNS)
    wb.create_sheet(SHEET_RECURRING).append(RECURRING_COLUMNS)

    return wb

def _collect_import_movements(expense_rows, income_rows):
    errors = []
    insert_expenses = []
    for idx, row in enumerate(expense_rows, start=2):
        notes_val = row.get("notes")
        notes = "" if notes_val is None else str(notes_val)
        try:
            amount = _parse_excel_expense_amount(row.get("amount"), SHEET_EXPENSES, idx)
        except ValueError as exc:
            errors.append(str(exc))
            continue
        cat_name = row.get("category_name")
        acc_name = row.get("account_name")
        if cat_name is None or not str(cat_name).strip():
            errors.append(f"{SHEET_EXPENSES} row {idx}: missing category_name")
            continue
        if acc_name is None or not str(acc_name).strip():
            errors.append(f"{SHEET_EXPENSES} row {idx}: missing account_name")
            continue
        cat_name = str(cat_name).strip()
        acc_name = str(acc_name).strip()
        try:
            spent_at = _parse_excel_movement_date(row.get("spent_at"), SHEET_EXPENSES, idx, "spent_at")
        except ValueError as exc:
            errors.append(str(exc))
            continue
        created_at = None
        try:
            if row.get("created_at") not in (None, ""):
                created_at = _optional_created_at(row.get("created_at"), SHEET_EXPENSES, idx)
        except ValueError as exc:
            errors.append(str(exc))
            continue
        insert_expenses.append((notes, amount, cat_name, acc_name, spent_at, created_at))

    insert_income = []
    for idx, row in enumerate(income_rows, start=2):
        notes_val = row.get("notes")
        notes = "" if notes_val is None else str(notes_val)
        try:
            amount = _parse_excel_income_amount(row.get("amount"), SHEET_INCOME, idx)
        except ValueError as exc:
            errors.append(str(exc))
            continue
        cat_name = row.get("category_name")
        acc_name = row.get("account_name")
        if cat_name is None or not str(cat_name).strip():
            errors.append(f"{SHEET_INCOME} row {idx}: missing category_name")
            continue
        if acc_name is None or not str(acc_name).strip():
            errors.append(f"{SHEET_INCOME} row {idx}: missing account_name")
            continue
        cat_name = str(cat_name).strip()
        acc_name = str(acc_name).strip()
        try:
            received_at = _parse_excel_movement_date(
                row.get("received_at"), SHEET_INCOME, idx, "received_at"
            )
        except ValueError as exc:
            errors.append(str(exc))
            continue
        created_at = None
        try:
            if row.get("created_at") not in (None, ""):
                created_at = _optional_created_at(row.get("created_at"), SHEET_INCOME, idx)
        except ValueError as exc:
            errors.append(str(exc))
            continue
        insert_income.append((notes, amount, cat_name, acc_name, received_at, created_at))

    return insert_expenses, insert_income, errors

def _required_text(row, key, sheet, row_num, errors):
    value = row.get(key)
    if value is None or not str(value).strip():
        errors.append(f"{sheet} row {row_num}: missing {key}")
        return None
    return str(value).strip()


def _parse_excel_number(row, key, sheet, row_num, errors, *, default=None, minimum=None):
    raw = row.get(key)
    if raw is None or str(raw).strip() == "":
        if default is None:
            errors.append(f"{sheet} row {row_num}: missing {key}")
        return default
    try:
        value = float(raw)
    except (TypeError, ValueError):
        errors.append(f"{sheet} row {row_num}: invalid {key}")
        return None
    if minimum is not None and value < minimum:
        errors.append(f"{sheet} row {row_num}: {key} must be >= {minimum}")
        return None
    return value


def _parse_tx_type(row, sheet, row_num, errors):
    value = str(row.get("tx_type") or "").strip().lower()
    if value not in ("buy", "sell"):
        errors.append(f"{sheet} row {row_num}: tx_type must be 'buy' or 'sell'")
        return None
    return value


def _optional_text(row, key):
    value = row.get(key)
    return "" if value is None else str(value).strip()


def _collect_import_transfers(rows):
    errors = []
    out = []
    for idx, row in enumerate(rows, start=2):
        from_name = _required_text(row, "from_account_name", SHEET_TRANSFERS, idx, errors)
        to_name = _required_text(row, "to_account_name", SHEET_TRANSFERS, idx, errors)
        amount = _parse_excel_number(row, "amount", SHEET_TRANSFERS, idx, errors, minimum=0)
        if from_name and to_name and from_name == to_name:
            errors.append(f"{SHEET_TRANSFERS} row {idx}: from and to accounts must differ")
        if amount is not None and amount <= 0:
            errors.append(f"{SHEET_TRANSFERS} row {idx}: amount must be positive")
        try:
            transferred_at = _parse_excel_movement_date(
                row.get("transferred_at"), SHEET_TRANSFERS, idx, "transferred_at"
            )
        except ValueError as exc:
            errors.append(str(exc))
            continue
        created_at = None
        if row.get("created_at") not in (None, ""):
            try:
                created_at = _optional_created_at(row.get("created_at"), SHEET_TRANSFERS, idx)
            except ValueError as exc:
                errors.append(str(exc))
                continue
        if from_name and to_name and amount:
            out.append(
                (from_name, to_name, abs(amount), transferred_at, _optional_text(row, "notes"), created_at)
            )
    return out, errors


def _collect_import_investments(rows, sheet, columns, id_fields):
    """Stocks and crypto differ only in their identifying/venue columns."""
    errors = []
    out = []
    for idx, row in enumerate(rows, start=2):
        ids = [_required_text(row, field, sheet, idx, errors) for field in id_fields]
        tx_type = _parse_tx_type(row, sheet, idx, errors)
        quantity = _parse_excel_number(row, "quantity", sheet, idx, errors, minimum=0)
        price = _parse_excel_number(row, "price_per_unit", sheet, idx, errors, minimum=0)
        fee = _parse_excel_number(row, "fee", sheet, idx, errors, default=0.0, minimum=0)
        if quantity is not None and quantity <= 0:
            errors.append(f"{sheet} row {idx}: quantity must be positive")
        try:
            transacted_at = _parse_excel_movement_date(
                row.get("transacted_at"), sheet, idx, "transacted_at"
            )
        except ValueError as exc:
            errors.append(str(exc))
            continue
        created_at = None
        if row.get("created_at") not in (None, ""):
            try:
                created_at = _optional_created_at(row.get("created_at"), sheet, idx)
            except ValueError as exc:
                errors.append(str(exc))
                continue
        venue_field = columns[7]  # "broker" for stocks, "exchange" for crypto
        if all(ids) and tx_type and quantity and price is not None:
            out.append(
                tuple(ids)
                + (
                    tx_type,
                    quantity,
                    price,
                    fee or 0.0,
                    _optional_text(row, venue_field),
                    transacted_at,
                    _optional_text(row, "notes"),
                    created_at,
                )
            )
    return out, errors


def _parse_applied_months(raw, sheet, row_num, errors):
    if raw is None or not str(raw).strip():
        return []
    months = []
    for chunk in str(raw).split(","):
        ym = chunk.strip()
        if not ym:
            continue
        if not re.fullmatch(r"\d{4}-\d{2}", ym):
            errors.append(f"{sheet} row {row_num}: invalid applied_months entry {ym!r}")
            continue
        months.append(ym)
    return months


def _collect_import_recurring(rows):
    errors = []
    out = []
    for idx, row in enumerate(rows, start=2):
        entry_type = str(row.get("entry_type") or "").strip().lower()
        if entry_type not in ("expense", "income"):
            errors.append(f"{SHEET_RECURRING} row {idx}: entry_type must be 'expense' or 'income'")
            continue
        amount = _parse_excel_number(row, "amount", SHEET_RECURRING, idx, errors, minimum=0)
        if amount is not None and amount <= 0:
            errors.append(f"{SHEET_RECURRING} row {idx}: amount must be positive")
        cat_name = _required_text(row, "category_name", SHEET_RECURRING, idx, errors)
        acc_name = _required_text(row, "account_name", SHEET_RECURRING, idx, errors)
        day = _parse_excel_number(row, "day_of_month", SHEET_RECURRING, idx, errors)
        if day is not None and not 1 <= int(day) <= 31:
            errors.append(f"{SHEET_RECURRING} row {idx}: day_of_month must be 1-31")
            day = None
        months_to_run = None
        if row.get("months_to_run") not in (None, ""):
            months_to_run = _parse_excel_number(
                row, "months_to_run", SHEET_RECURRING, idx, errors, minimum=1
            )
            if months_to_run is not None:
                months_to_run = int(months_to_run)
        enabled_raw = row.get("enabled")
        enabled = 1 if enabled_raw in (None, "") else int(bool(float(enabled_raw)))
        created_at = None
        if row.get("created_at") not in (None, ""):
            try:
                created_at = _optional_created_at(row.get("created_at"), SHEET_RECURRING, idx)
            except ValueError as exc:
                errors.append(str(exc))
                continue
        applied = _parse_applied_months(row.get("applied_months"), SHEET_RECURRING, idx, errors)
        if amount and cat_name and acc_name and day is not None:
            out.append(
                (
                    entry_type,
                    abs(amount),
                    cat_name,
                    acc_name,
                    int(day),
                    months_to_run,
                    _optional_text(row, "notes"),
                    enabled,
                    created_at,
                    applied,
                )
            )
    return out, errors


def _run_import_workbook(wb, replace_movements, sync_opening_balances, user_id):
    errors = []
    required_sheets = {
        SHEET_ACCOUNTS,
        SHEET_EXPENSE_CATEGORIES,
        SHEET_INCOME_CATEGORIES,
        SHEET_EXPENSES,
        SHEET_INCOME,
    }
    missing = [name for name in required_sheets if name not in wb.sheetnames]
    if missing:
        return [f"Missing sheets: {', '.join(missing)}"]

    accounts_rows = _sheet_as_dicts(wb[SHEET_ACCOUNTS])
    expense_cat_rows = _sheet_as_dicts(wb[SHEET_EXPENSE_CATEGORIES])
    income_cat_rows = _sheet_as_dicts(wb[SHEET_INCOME_CATEGORIES])
    expense_rows = _sheet_as_dicts(wb[SHEET_EXPENSES])
    income_rows = _sheet_as_dicts(wb[SHEET_INCOME])

    insert_expenses, insert_income, parse_errors = _collect_import_movements(
        expense_rows, income_rows
    )

    # Sheets added in format version 2. Absent in version 1 workbooks, so each
    # one is optional and simply yields nothing when the sheet is not there.
    def _rows_of(sheet_name):
        return _sheet_as_dicts(wb[sheet_name]) if sheet_name in wb.sheetnames else []

    insert_transfers, transfer_errors = _collect_import_transfers(_rows_of(SHEET_TRANSFERS))
    insert_stocks, stock_errors = _collect_import_investments(
        _rows_of(SHEET_STOCKS), SHEET_STOCKS, STOCK_COLUMNS, ("symbol", "ticker", "instrument_name")
    )
    insert_crypto, crypto_errors = _collect_import_investments(
        _rows_of(SHEET_CRYPTO), SHEET_CRYPTO, CRYPTO_COLUMNS, ("coin_id", "coin_symbol", "coin_name")
    )
    insert_recurring, recurring_errors = _collect_import_recurring(_rows_of(SHEET_RECURRING))

    parse_errors = (
        parse_errors + transfer_errors + stock_errors + crypto_errors + recurring_errors
    )
    if parse_errors:
        return parse_errors

    expense_cats_from_movements = {row[2] for row in insert_expenses} | {
        row[2] for row in insert_recurring if row[0] == "expense"
    }
    income_cats_from_movements = {row[2] for row in insert_income} | {
        row[2] for row in insert_recurring if row[0] == "income"
    }
    accounts_from_movements = (
        {row[3] for row in insert_expenses}
        | {row[3] for row in insert_income}
        | {row[0] for row in insert_transfers}
        | {row[1] for row in insert_transfers}
        | {row[3] for row in insert_recurring}
    )

    conn = get_connection()
    uid = int(user_id)
    try:
        conn.execute("BEGIN")

        for idx, row in enumerate(accounts_rows, start=2):
            name = row.get("name")
            if name is None or not str(name).strip():
                errors.append(f"{SHEET_ACCOUNTS} row {idx}: missing name")
                continue
            name = str(name).strip()
            opening_raw = row.get("opening_balance")
            if opening_raw is None or str(opening_raw).strip() == "":
                opening_balance = 0.0
            else:
                try:
                    opening_balance = float(opening_raw)
                except (TypeError, ValueError):
                    errors.append(f"{SHEET_ACCOUNTS} row {idx}: invalid opening_balance")
                    continue
            conn.execute(
                "INSERT OR IGNORE INTO accounts(user_id, name, opening_balance) VALUES (?, ?, ?)",
                (uid, name, opening_balance),
            )
            if sync_opening_balances:
                conn.execute(
                    "UPDATE accounts SET opening_balance = ? WHERE user_id = ? AND name = ?",
                    (opening_balance, uid, name),
                )

        if errors:
            conn.rollback()
            return errors

        for idx, row in enumerate(expense_cat_rows, start=2):
            name = row.get("name")
            if name is None or not str(name).strip():
                continue
            conn.execute(
                "INSERT OR IGNORE INTO categories(user_id, name) VALUES (?, ?)",
                (uid, str(name).strip()),
            )

        for name in sorted(expense_cats_from_movements):
            conn.execute(
                "INSERT OR IGNORE INTO categories(user_id, name) VALUES (?, ?)",
                (uid, name),
            )

        for idx, row in enumerate(income_cat_rows, start=2):
            name = row.get("name")
            if name is None or not str(name).strip():
                continue
            conn.execute(
                "INSERT OR IGNORE INTO income_categories(user_id, name) VALUES (?, ?)",
                (uid, str(name).strip()),
            )

        for name in sorted(income_cats_from_movements):
            conn.execute(
                "INSERT OR IGNORE INTO income_categories(user_id, name) VALUES (?, ?)",
                (uid, name),
            )

        for acc_name in sorted(accounts_from_movements):
            conn.execute(
                "INSERT OR IGNORE INTO accounts(user_id, name, opening_balance) VALUES (?, ?, ?)",
                (uid, acc_name, 0.0),
            )

        if replace_movements:
            conn.execute("DELETE FROM expenses WHERE user_id = ?", (uid,))
            conn.execute("DELETE FROM income_entries WHERE user_id = ?", (uid,))
            conn.execute("DELETE FROM account_transfers WHERE user_id = ?", (uid,))
            conn.execute("DELETE FROM stock_transactions WHERE user_id = ?", (uid,))
            conn.execute("DELETE FROM crypto_transactions WHERE user_id = ?", (uid,))
            conn.execute(
                """
                DELETE FROM recurring_applied
                WHERE recurring_id IN (SELECT id FROM recurring_entries WHERE user_id = ?)
                """,
                (uid,),
            )
            conn.execute("DELETE FROM recurring_entries WHERE user_id = ?", (uid,))

        for notes, amount, cat_name, acc_name, spent_at, created_at in insert_expenses:
            category_id = _lookup_category_id(conn, cat_name, uid, expense=True)
            account_id = _lookup_account_id(conn, acc_name, uid)
            if category_id is None:
                errors.append(f"{SHEET_EXPENSES}: unknown expense category {cat_name!r}")
            if account_id is None:
                errors.append(f"{SHEET_EXPENSES}: unknown account {acc_name!r}")

        for notes, amount, cat_name, acc_name, received_at, created_at in insert_income:
            category_id = _lookup_category_id(conn, cat_name, uid, expense=False)
            account_id = _lookup_account_id(conn, acc_name, uid)
            if category_id is None:
                errors.append(f"{SHEET_INCOME}: unknown income category {cat_name!r}")
            if account_id is None:
                errors.append(f"{SHEET_INCOME}: unknown account {acc_name!r}")

        if errors:
            conn.rollback()
            return errors

        for notes, amount, cat_name, acc_name, spent_at, created_at in insert_expenses:
            category_id = _lookup_category_id(conn, cat_name, uid, expense=True)
            account_id = _lookup_account_id(conn, acc_name, uid)
            if created_at:
                conn.execute(
                    """
                    INSERT INTO expenses (user_id, notes, amount, category_id, account_id, spent_at, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (uid, notes, amount, category_id, account_id, spent_at, created_at),
                )
            else:
                conn.execute(
                    """
                    INSERT INTO expenses (user_id, notes, amount, category_id, account_id, spent_at)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (uid, notes, amount, category_id, account_id, spent_at),
                )

        for notes, amount, cat_name, acc_name, received_at, created_at in insert_income:
            category_id = _lookup_category_id(conn, cat_name, uid, expense=False)
            account_id = _lookup_account_id(conn, acc_name, uid)
            if created_at:
                conn.execute(
                    """
                    INSERT INTO income_entries (user_id, notes, amount, category_id, account_id, received_at, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (uid, notes, amount, category_id, account_id, received_at, created_at),
                )
            else:
                conn.execute(
                    """
                    INSERT INTO income_entries (user_id, notes, amount, category_id, account_id, received_at)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (uid, notes, amount, category_id, account_id, received_at),
                )

        for from_name, to_name, amount, transferred_at, notes, created_at in insert_transfers:
            from_id = _lookup_account_id(conn, from_name, uid)
            to_id = _lookup_account_id(conn, to_name, uid)
            if from_id is None or to_id is None:
                errors.append(
                    f"{SHEET_TRANSFERS}: unknown account "
                    f"{from_name if from_id is None else to_name!r}"
                )
                continue
            if created_at:
                conn.execute(
                    """
                    INSERT INTO account_transfers
                        (user_id, from_account_id, to_account_id, amount, transferred_at, notes, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (uid, from_id, to_id, amount, transferred_at, notes, created_at),
                )
            else:
                conn.execute(
                    """
                    INSERT INTO account_transfers
                        (user_id, from_account_id, to_account_id, amount, transferred_at, notes)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (uid, from_id, to_id, amount, transferred_at, notes),
                )

        for symbol, ticker, name, tx_type, qty, price, fee, broker, at, notes, created_at in insert_stocks:
            conn.execute(
                """
                INSERT INTO stock_transactions
                    (user_id, symbol, ticker, instrument_name, tx_type, quantity,
                     price_per_unit, fee, broker, transacted_at, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (uid, symbol, ticker, name, tx_type, qty, price, fee, broker, at, notes),
            )

        for cid, csym, cname, tx_type, qty, price, fee, exchange, at, notes, created_at in insert_crypto:
            conn.execute(
                """
                INSERT INTO crypto_transactions
                    (user_id, coin_id, coin_symbol, coin_name, tx_type, quantity,
                     price_per_unit, fee, exchange, transacted_at, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (uid, cid, csym, cname, tx_type, qty, price, fee, exchange, at, notes),
            )

        for (
            entry_type, amount, cat_name, acc_name, day, months_to_run,
            notes, enabled, created_at, applied_months,
        ) in insert_recurring:
            category_id = _lookup_category_id(conn, cat_name, uid, expense=entry_type == "expense")
            account_id = _lookup_account_id(conn, acc_name, uid)
            if category_id is None or account_id is None:
                errors.append(
                    f"{SHEET_RECURRING}: unknown "
                    + (f"category {cat_name!r}" if category_id is None else f"account {acc_name!r}")
                )
                continue
            if created_at:
                cursor = conn.execute(
                    """
                    INSERT INTO recurring_entries
                        (user_id, entry_type, amount, category_id, account_id, day_of_month,
                         months_to_run, notes, enabled, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (uid, entry_type, amount, category_id, account_id, day,
                     months_to_run, notes, enabled, created_at),
                )
            else:
                cursor = conn.execute(
                    """
                    INSERT INTO recurring_entries
                        (user_id, entry_type, amount, category_id, account_id, day_of_month,
                         months_to_run, notes, enabled)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (uid, entry_type, amount, category_id, account_id, day,
                     months_to_run, notes, enabled),
                )
            rule_id = cursor.lastrowid
            for ym in applied_months:
                conn.execute(
                    "INSERT OR IGNORE INTO recurring_applied (recurring_id, ym) VALUES (?, ?)",
                    (rule_id, ym),
                )

        if errors:
            conn.rollback()
            return errors

        conn.commit()
        return []
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
