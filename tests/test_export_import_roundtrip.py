"""Export -> import round trip.

Version 1 of the export only covered accounts, categories, expenses and income,
so transfers, investments and recurring rules were silently dropped by a
backup/restore cycle - and because transfers move money between accounts, the
restored balances were wrong too. Version 2 adds those sheets.
"""

import io
from datetime import datetime

from openpyxl import Workbook, load_workbook

import app as vb_app
import config
import excel_io

from .conftest import csrf_token, register_and_login


def _ids_for(username):
    conn = vb_app.get_connection()
    uid = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()["id"]
    account_id = conn.execute(
        "SELECT id FROM accounts WHERE user_id = ? ORDER BY id LIMIT 1", (uid,)
    ).fetchone()["id"]
    category_id = conn.execute(
        "SELECT id FROM categories WHERE user_id = ? ORDER BY id LIMIT 1", (uid,)
    ).fetchone()["id"]
    conn.close()
    return uid, account_id, category_id


def _seed_everything(client, username):
    """One row of every kind the app stores."""
    uid, account_id, category_id = _ids_for(username)
    token = csrf_token(client)

    client.post(
        "/accounts/add",
        data={"name": "Savings", "opening_balance": "250.00", "_csrf_token": token},
        follow_redirects=True,
    )
    conn = vb_app.get_connection()
    savings_id = conn.execute(
        "SELECT id FROM accounts WHERE user_id = ? AND name = 'Savings'", (uid,)
    ).fetchone()["id"]
    income_cat_id = conn.execute(
        "SELECT id FROM income_categories WHERE user_id = ? ORDER BY id LIMIT 1", (uid,)
    ).fetchone()["id"]
    conn.close()

    client.post(
        "/expenses/add",
        data={
            "notes": "groceries", "amount": "-42.50",
            "category_id": str(category_id), "account_id": str(account_id),
            "_csrf_token": csrf_token(client),
        },
        follow_redirects=True,
    )
    client.post(
        "/income/add",
        data={
            "notes": "paycheck", "amount": "2000",
            "category_id": str(income_cat_id), "account_id": str(account_id),
            "_csrf_token": csrf_token(client),
        },
        follow_redirects=True,
    )
    client.post(
        "/transfers/add",
        data={
            "from_account_id": str(account_id), "to_account_id": str(savings_id),
            "amount": "300", "transferred_at": "2026-05-04", "notes": "to savings",
            "_csrf_token": csrf_token(client),
        },
        follow_redirects=True,
    )
    client.post(
        "/stocks/add",
        data={
            "symbol": "VUAA.DE", "ticker": "VUAA", "instrument_name": "Vanguard S&P 500",
            "tx_type": "buy", "quantity": "4", "price_per_unit": "92.15", "fee": "1.20",
            "broker": "Freedom24", "transacted_at": "2026-02-11", "notes": "monthly buy",
            "_csrf_token": csrf_token(client),
        },
        follow_redirects=True,
    )
    client.post(
        "/crypto/add",
        data={
            "coin_id": "bitcoin", "coin_symbol": "BTC", "coin_name": "Bitcoin",
            "tx_type": "buy", "quantity": "0.125", "price_per_unit": "41000", "fee": "5",
            "exchange": "Kraken", "transacted_at": "2026-03-09", "notes": "dca",
            "_csrf_token": csrf_token(client),
        },
        follow_redirects=True,
    )
    # Due today: a rule only posts for months where created_at <= due <= today,
    # so any other day of the month would leave applied_months empty.
    client.post(
        "/recurring/add",
        data={
            "category_choice": f"e-{category_id}", "amount": "800",
            "account_id": str(account_id), "day_of_month": str(datetime.now().day),
            "indefinitely": "1", "notes": "rent",
            "_csrf_token": csrf_token(client),
        },
        follow_redirects=True,
    )
    # Loading the dashboard posts the recurring rule for the due months, which
    # is what fills recurring_applied.
    client.get("/")


def _snapshot(uid):
    conn = vb_app.get_connection()
    out = {
        "accounts": [
            (r["name"], r["opening_balance"])
            for r in conn.execute(
                "SELECT name, opening_balance FROM accounts WHERE user_id = ? ORDER BY name", (uid,)
            )
        ],
        "transfers": [
            (r["amount"], r["transferred_at"], r["notes"])
            for r in conn.execute(
                "SELECT amount, transferred_at, notes FROM account_transfers WHERE user_id = ? ORDER BY id",
                (uid,),
            )
        ],
        "stocks": [
            (r["symbol"], r["tx_type"], r["quantity"], r["price_per_unit"], r["fee"], r["broker"])
            for r in conn.execute(
                "SELECT symbol, tx_type, quantity, price_per_unit, fee, broker FROM stock_transactions WHERE user_id = ? ORDER BY id",
                (uid,),
            )
        ],
        "crypto": [
            (r["coin_id"], r["tx_type"], r["quantity"], r["price_per_unit"], r["exchange"])
            for r in conn.execute(
                "SELECT coin_id, tx_type, quantity, price_per_unit, exchange FROM crypto_transactions WHERE user_id = ? ORDER BY id",
                (uid,),
            )
        ],
        "recurring": [
            (r["entry_type"], r["amount"], r["day_of_month"], r["notes"], r["enabled"])
            for r in conn.execute(
                "SELECT entry_type, amount, day_of_month, notes, enabled FROM recurring_entries WHERE user_id = ? ORDER BY id",
                (uid,),
            )
        ],
        "applied_months": sorted(
            r["ym"]
            for r in conn.execute(
                """
                SELECT ra.ym FROM recurring_applied ra
                JOIN recurring_entries r ON r.id = ra.recurring_id
                WHERE r.user_id = ?
                """,
                (uid,),
            )
        ),
        "expense_count": conn.execute(
            "SELECT COUNT(*) n FROM expenses WHERE user_id = ?", (uid,)
        ).fetchone()["n"],
    }
    conn.close()
    return out


def test_export_covers_every_sheet(client):
    register_and_login(client, "exportall")
    _seed_everything(client, "exportall")

    wb = load_workbook(io.BytesIO(client.get("/export/excel").data))
    for sheet in (
        config.SHEET_ACCOUNTS, config.SHEET_EXPENSE_CATEGORIES, config.SHEET_INCOME_CATEGORIES,
        config.SHEET_EXPENSES, config.SHEET_INCOME, config.SHEET_TRANSFERS,
        config.SHEET_STOCKS, config.SHEET_CRYPTO, config.SHEET_RECURRING,
    ):
        assert sheet in wb.sheetnames, f"export is missing the {sheet} sheet"

    # every added sheet actually carries its row, not just a header
    for sheet in (config.SHEET_TRANSFERS, config.SHEET_STOCKS, config.SHEET_CRYPTO, config.SHEET_RECURRING):
        assert wb[sheet].max_row >= 2, f"{sheet} exported no rows"

    applied = wb[config.SHEET_RECURRING].cell(row=2, column=len(excel_io.RECURRING_COLUMNS)).value
    assert applied, "recurring rule exported without its applied_months history"


def test_export_import_round_trip_preserves_everything(client, app):
    register_and_login(client, "roundtripsrc")
    _seed_everything(client, "roundtripsrc")
    src_uid, _, _ = _ids_for("roundtripsrc")
    before = _snapshot(src_uid)
    workbook_bytes = client.get("/export/excel").data

    other = app.test_client()
    register_and_login(other, "roundtripdst")
    dst_uid, _, _ = _ids_for("roundtripdst")

    resp = other.post(
        "/import/excel",
        data={
            "file": (io.BytesIO(workbook_bytes), "budget-export.xlsx"),
            "replace_movements": "1",
            "sync_opening_balances": "1",
            "_csrf_token": csrf_token(other),
        },
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert resp.status_code == 200
    assert b"Import failed" not in resp.data

    after = _snapshot(dst_uid)
    for key in ("accounts", "transfers", "stocks", "crypto", "recurring", "applied_months"):
        assert after[key] == before[key], f"{key} did not survive the round trip"


def test_restored_recurring_rule_does_not_repost(client, app):
    """The applied-month history has to come across, or the next dashboard load
    re-posts every month back to the rule's creation date."""
    register_and_login(client, "repostsrc")
    _seed_everything(client, "repostsrc")
    workbook_bytes = client.get("/export/excel").data

    other = app.test_client()
    register_and_login(other, "repostdst")
    dst_uid, _, _ = _ids_for("repostdst")
    other.post(
        "/import/excel",
        data={
            "file": (io.BytesIO(workbook_bytes), "budget-export.xlsx"),
            "replace_movements": "1",
            "sync_opening_balances": "1",
            "_csrf_token": csrf_token(other),
        },
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    right_after_import = _snapshot(dst_uid)["expense_count"]

    other.get("/")  # triggers apply_recurring_entries
    assert _snapshot(dst_uid)["expense_count"] == right_after_import


def test_version_1_workbook_still_imports(client):
    """Older exports have none of the sheets added in version 2."""
    register_and_login(client, "v1import")
    uid, _, _ = _ids_for("v1import")

    wb = Workbook()
    meta = wb.active
    meta.title = config.SHEET_META
    meta.append(["key", "value"])
    meta.append(["format_version", 1])
    wb.create_sheet(config.SHEET_ACCOUNTS).append(["name", "opening_balance"])
    wb[config.SHEET_ACCOUNTS].append(["Main", 0])
    wb.create_sheet(config.SHEET_EXPENSE_CATEGORIES).append(["name"])
    wb[config.SHEET_EXPENSE_CATEGORIES].append(["General"])
    wb.create_sheet(config.SHEET_INCOME_CATEGORIES).append(["name"])
    wb[config.SHEET_INCOME_CATEGORIES].append(["General"])
    wb.create_sheet(config.SHEET_EXPENSES).append(
        ["notes", "amount", "category_name", "account_name", "spent_at", "created_at"]
    )
    wb[config.SHEET_EXPENSES].append(["old expense", -10.0, "General", "Main", "2026-01-05", None])
    wb.create_sheet(config.SHEET_INCOME).append(
        ["notes", "amount", "category_name", "account_name", "received_at", "created_at"]
    )

    buf = io.BytesIO()
    wb.save(buf)
    resp = client.post(
        "/import/excel",
        data={
            "file": (io.BytesIO(buf.getvalue()), "old-export.xlsx"),
            "_csrf_token": csrf_token(client),
        },
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert resp.status_code == 200
    assert b"Import failed" not in resp.data

    conn = vb_app.get_connection()
    row = conn.execute(
        "SELECT amount FROM expenses WHERE user_id = ? AND notes = 'old expense'", (uid,)
    ).fetchone()
    conn.close()
    assert row is not None and row["amount"] == -10.0
